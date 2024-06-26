'''
┌─┐  ┌─┐  ┌┬┐  ┬  ┌┐┌  ┬
│ ┬  ├┤   │││  │  │││  │
└─┘  └─┘  ┴ ┴  ┴  ┘└┘  ┴ plot
---------------------------------------------------
-*- coding: utf-8 -*-                              |
title            : geminiphage.py                  |
description      : gemini phage functions          |
author           : dooguypapua                     |
lastmodification : 20210720                        |
version          : 0.1                             |
python_version   : 3.8.5                           |
---------------------------------------------------
'''

import sys
import re
import os
import shutil
import xlsxwriter
import xml.etree.ElementTree as ET
from tqdm import tqdm
from typing import Tuple
from yaspin import yaspin
from xvfbwrapper import Xvfb
from yaspin.spinners import Spinners
import matplotlib.pyplot as plt
from collections import OrderedDict
from dna_features_viewer import GraphicFeature, GraphicRecord
from svgpathtools import svg2paths
from cairosvg import svg2png
from openpyxl import Workbook, load_workbook, styles
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment
import geminiset
from geminini import fct_checker, get_input_files, printcolor, path_converter, load_json
from geminini import random_hex_color, linear_gradient, title, exit_gemini, cat_lstfiles, to_ranges
from geminini import requires_white_text, shift_dico_pos, read_file, get_gemini_path
from geminicluster import mmseqs_rbh
from geminiparse import make_gff_dict, make_gbk_dict, gbk_to_faa, make_fasta_dict


@fct_checker
def gff_to_linear_geneplot(pathIN: str, pathOUT: str, pathLT: str = "None", length: int = -1, ext: str = ".gff") -> Tuple[str, str, str, int, str]:
    '''
     ------------------------------------------------------------
    |                  GFF3 TO LINEAR GENE PLOT                  |
    |------------------------------------------------------------|
    |           Create linear gene plot from GFF3 file           |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN : path of input files or folder (required)       |
    |    pathOUT: path of output files (required)                |
    |    pathLT : path of input LT list file (default=None)      |
    |    length : representation length (default=-1 use sequence)|
    |    ext    : extension of input files (default=.gff)        |
     ------------------------------------------------------------
    '''
    lstFiles, maxpathSize = get_input_files(pathIN, "gff_to_linear_geneplot", [ext])
    if len(lstFiles) == 0:
        printcolor("[ERROR: gff_to_linear_geneplot]\nAny input files found\n", 1, "212;64;89", "None", True)
        exit_gemini()
    pathOUT = path_converter(pathOUT)
    os.makedirs(pathOUT, exist_ok=True)
    # Retrieve LT list if required
    if pathLT != "None":
        setLT = set()
        pathLT = path_converter(pathLT)
        LTFILE = open(pathLT, 'r')
        lstLines = LTFILE.read().split("\n")
        LTFILE.close()
        for line in lstLines:
            if line != "" and not line[0] == "#":
                setLT.add(line)
    # Browse GFF3
    printcolor("♊ Plotting"+"\n")
    pbar = tqdm(total=int(len(lstFiles)), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    for pathGFF in lstFiles:
        file = os.path.basename(pathGFF)
        orgName = file.replace(ext, "").replace("."+ext, "")
        pathPNG = pathOUT+"/"+orgName+".png"
        pathSVG = pathOUT+"/"+orgName+".svg"
        pbar.set_description_str(orgName+" ".rjust(maxpathSize-len(orgName)))
        features = []
        # ***** PARSE GFF ***** #
        dicoGFF = make_gff_dict(pathIN=pathGFF, ext=ext)
        startRegion = 0
        endRegion = 0
        # ***** BROWSE GENES ***** #
        for geneType in dicoGFF[orgName]:
            if geneType != 'length':
                for geneEntry in dicoGFF[orgName][geneType]:
                    if geneType == "CDS":
                        if geneEntry['strand'] == "+": color = "#2a7fff"
                        else: color = "#dddfff"
                        # if geneEntry['attributes']['locus_tag'] in ["VP6E351A_0009","VP6E351A_0101","VP6E351A_0201"]: color = "red"
                    elif geneType == "tRNA":
                        color = "#37c8ab"
                    else:
                        continue
                    if 'locus_tag' in geneEntry['attributes']:
                        if geneType == "CDS":
                            geneFeature = GraphicFeature(start=geneEntry['start'], end=geneEntry['end'], strand=int(geneEntry['strand']+"1"), color=color, linewidth=1) #, label=geneEntry['attributes']['locus_tag'].split(" ")[0])
                        else:
                            geneFeature = GraphicFeature(start=geneEntry['start'], end=geneEntry['end'], strand=int(geneEntry['strand']+"1"), color=color, linewidth=1, label=geneEntry['attributes']['product'].split("#")[0]+"-"+geneEntry['attributes']['product'].split("#")[2])                            
                    else:
                        geneFeature = GraphicFeature(start=geneEntry['start'], end=geneEntry['end'], strand=int(geneEntry['strand']+"1"), color=color, linewidth=1)
                    # if pathLT == "None":
                    #     features.append(geneFeature)
                    # elif geneEntry['attributes']['locus_tag'] in setLT:
                    #     if startRegion == 0:
                    #         startRegion = geneEntry['start']
                    #     endRegion = geneEntry['end']
                    features.append(geneFeature)
        # ***** PLOT GENES ***** #
        if length != -1:
            seqLen = length
        elif pathLT == "None":
            seqLen = dicoGFF[orgName]['length']
        else:
            seqLen = endRegion-startRegion
        record = GraphicRecord(sequence_length=seqLen+int(seqLen/10), features=features, first_index=startRegion-100)
        ax, _ = record.plot(figure_width=50)
        ax.figure.savefig(pathPNG, dpi=300)
        ax.figure.savefig(pathSVG)
        plt.close('all')
        pbar.update(1)
        title("Plotting", pbar)
    pbar.close()


@fct_checker
def rbh_linear_plot(pathIN: str, pathCLUSTER: str, pathOUT: str, distinctColor: bool = False, pathSUBCORE: str = "None", ext: str = ".gff") -> Tuple[str, str, str, bool, str, str]:
    '''
     ------------------------------------------------------------
    |              RBH CLUSTERS TO LINEAR GENE PLOTS             |
    |------------------------------------------------------------|
    |      Create linear gene plots from mutliples GFF3 file     |
    |             and colorized based on rbh cluster             |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN       : path of input files or folder (required) |
    |    pathCLUSTER  : path of JSON rbh cluster (required)      |
    |    pathOUT      : path of output files (required)          |
    |    distinctColor: distinct families colors (default=False) |
    |    pathSUBCORE  : subcore organism list file (default=None)|
    |    ext          : extension of input files (default=.gff)  |
     ------------------------------------------------------------
    '''

    dicoBlast = {0: {'VPD486_0300___Vibrio_phage_D486', 'VPK566_0300___Vibrio_phage_K566', 'P001501635_CDS_0417___Vibrio_phage_ValKK3', 'VPK571_0307___Vibrio_phage_K571', 'VPK491_0307___Vibrio_phage_K491', 'VPF86_0221___Vibrio_phage_F86', 'VPK575_0306___Vibrio_phage_K575', 'P004771355_CDS_0279___Vibrio_phage_Va3', 'VPD483_0307___Vibrio_phage_D483', 'VPD481_0303___Vibrio_phage_D481', 'VPD527_0303___Vibrio_phage_D527', 'P000843785_CDS_0262___Vibrio_phage_KVP40', 'VPD484_0300___Vibrio_phage_D484', 'P020809685_CDS_0251___Vibrio_phage_VS1', 'P022984185_CDS_0152___Vibrio_phage_XZ1', 'P024234235_CDS_0092___Vibrio_phage_PVA23', 'P000914495_CDS_0165___Vibrio_phage_VH7D', 'VPK567_0302___Vibrio_phage_K567', 'VPD525_0307___Vibrio_phage_D525', 'P000910195_CDS_0141___Vibrio_phage_nt-1', 'VP6E351A_0309___Vibrio_phage_6E35-1a', 'P003927315_CDS_0269___Vibrio_phage_1.081.O._10N.286.52.C2', 'VPD490_0307___Vibrio_phage_D490', 'P012361135_CDS_0334___Vibrio_phage_V09', 'P030927055_CDS_0077___Vibrio_phage_PC-Liy1', 'P002630725_CDS_0257___Vibrio_phage_phi-pp2', 'VPD485_0305___Vibrio_phage_D485', 'P024381305_CDS_0077___Vibrio_phage_PVA8', 'P009931825_CDS_0262___Vibrio_phage_VH1_2019', 'VPD482_0309___Vibrio_phage_D482', 'VPD491S_0303___Vibrio_phage_D491', 'P014337555_CDS_0321___Vibrio_phage_vB_ValM_R11Z', 'P027885675_CDS_0077___Vibrio_phage_vB_ValM_PVA8', 'P002757135_CDS_0141___Vibrio_phage_phi-ST2', 'P029686055_CDS_0383___Vibrio_phage_vB_VspM_VS2', 'VPD480_0303___Vibrio_phage_D480', 'P002757115_CDS_0420___Vibrio_phage_phi-Grn1', 'P014337545_CDS_0321___Vibrio_phage_vB_ValM_R10Z'}, 1: {'P003927315_CDS_0270___Vibrio_phage_1.081.O._10N.286.52.C2'}, 2: {'P002757135_CDS_0143___Vibrio_phage_phi-ST2', 'P014337545_CDS_0319___Vibrio_phage_vB_ValM_R10Z', 'P000914495_CDS_0167___Vibrio_phage_VH7D', 'P002757115_CDS_0422___Vibrio_phage_phi-Grn1', 'P022984185_CDS_0150___Vibrio_phage_XZ1', 'P004771355_CDS_0277___Vibrio_phage_Va3', 'P003927315_CDS_0277___Vibrio_phage_1.081.O._10N.286.52.C2', 'P014337555_CDS_0319___Vibrio_phage_vB_ValM_R11Z', 'P001501635_CDS_0415___Vibrio_phage_ValKK3', 'P003927315_CDS_0272___Vibrio_phage_1.081.O._10N.286.52.C2', 'P000910195_CDS_0133___Vibrio_phage_nt-1', 'P003927315_CDS_0271___Vibrio_phage_1.081.O._10N.286.52.C2', 'P024234235_CDS_0094___Vibrio_phage_PVA23'}, 3: {'P003927315_CDS_0273___Vibrio_phage_1.081.O._10N.286.52.C2'}, 4: {'P003927315_CDS_0274___Vibrio_phage_1.081.O._10N.286.52.C2'}, 5: {'P003927315_CDS_0275___Vibrio_phage_1.081.O._10N.286.52.C2'}, 6: {'P003927315_CDS_0276___Vibrio_phage_1.081.O._10N.286.52.C2', 'P000910195_CDS_0145___Vibrio_phage_nt-1'}, 7: {'P003927315_CDS_0278___Vibrio_phage_1.081.O._10N.286.52.C2'}, 8: {'P003927315_CDS_0279___Vibrio_phage_1.081.O._10N.286.52.C2'}, 9: {'P003927315_CDS_0280___Vibrio_phage_1.081.O._10N.286.52.C2'}, 10: {'P003927315_CDS_0281___Vibrio_phage_1.081.O._10N.286.52.C2'}, 11: {'P003927315_CDS_0282___Vibrio_phage_1.081.O._10N.286.52.C2'}, 12: {'VPD485_0314___Vibrio_phage_D485', 'P009931825_CDS_0269___Vibrio_phage_VH1_2019', 'P002630725_CDS_0264___Vibrio_phage_phi-pp2', 'VPD483_0326___Vibrio_phage_D483', 'P002630725_CDS_0265___Vibrio_phage_phi-pp2', 'VPK571_0316___Vibrio_phage_K571', 'P002757115_CDS_0424___Vibrio_phage_phi-Grn1', 'P014337555_CDS_0317___Vibrio_phage_vB_ValM_R11Z', 'P024381305_CDS_0092___Vibrio_phage_PVA8', 'VPD490_0325___Vibrio_phage_D490', 'VPD491S_0312___Vibrio_phage_D491', 'P012361135_CDS_0327___Vibrio_phage_V09', 'P009931825_CDS_0276___Vibrio_phage_VH1_2019', 'P027885675_CDS_0085___Vibrio_phage_vB_ValM_PVA8', 'P022984185_CDS_0148___Vibrio_phage_XZ1', 'P001501635_CDS_0413___Vibrio_phage_ValKK3', 'VPK491_0326___Vibrio_phage_K491', 'VPD481_0312___Vibrio_phage_D481', 'VPD484_0309___Vibrio_phage_D484', 'VPF86_0238___Vibrio_phage_F86', 'P014337545_CDS_0317___Vibrio_phage_vB_ValM_R10Z', 'VPD480_0322___Vibrio_phage_D480', 'VPD490_0316___Vibrio_phage_D490', 'P000914495_CDS_0169___Vibrio_phage_VH7D', 'P024234235_CDS_0096___Vibrio_phage_PVA23', 'VPD484_0319___Vibrio_phage_D484', 'P002757135_CDS_0145___Vibrio_phage_phi-ST2', 'VPD486_0309___Vibrio_phage_D486', 'VPD482_0327___Vibrio_phage_D482', 'VPK575_0315___Vibrio_phage_K575', 'VPD480_0312___Vibrio_phage_D480', 'P000910195_CDS_0134___Vibrio_phage_nt-1', 'VPK571_0326___Vibrio_phage_K571', 'P004771355_CDS_0275___Vibrio_phage_Va3', 'VPD527_0321___Vibrio_phage_D527', 'VPD527_0312___Vibrio_phage_D527', 'VPK491_0316___Vibrio_phage_K491', 'P002630725_CDS_0271___Vibrio_phage_phi-pp2', 'VPD485_0324___Vibrio_phage_D485', 'P000843785_CDS_0269___Vibrio_phage_KVP40', 'P030927055_CDS_0085___Vibrio_phage_PC-Liy1', 'P029686055_CDS_0396___Vibrio_phage_vB_VspM_VS2', 'VPK566_0315___Vibrio_phage_K566', 'P030927055_CDS_0092___Vibrio_phage_PC-Liy1', 'VPK567_0317___Vibrio_phage_K567', 'VP6E351A_0328___Vibrio_phage_6E35-1a', 'VPD481_0321___Vibrio_phage_D481', 'P029686055_CDS_0390___Vibrio_phage_vB_VspM_VS2', 'VPF86_0230___Vibrio_phage_F86', 'VP6E351A_0318___Vibrio_phage_6E35-1a', 'P024381305_CDS_0085___Vibrio_phage_PVA8', 'P027885675_CDS_0092___Vibrio_phage_vB_ValM_PVA8', 'VPD525_0327___Vibrio_phage_D525', 'VPD491S_0321___Vibrio_phage_D491', 'VPD482_0318___Vibrio_phage_D482', 'P020809685_CDS_0235___Vibrio_phage_VS1', 'VPD486_0318___Vibrio_phage_D486', 'VPD483_0316___Vibrio_phage_D483', 'P003927315_CDS_0283___Vibrio_phage_1.081.O._10N.286.52.C2', 'P012361135_CDS_0320___Vibrio_phage_V09', 'VPD525_0316___Vibrio_phage_D525', 'VPK575_0325___Vibrio_phage_K575', 'P000843785_CDS_0276___Vibrio_phage_KVP40', 'P020809685_CDS_0242___Vibrio_phage_VS1'}, 13: {'P022984185_CDS_0151___Vibrio_phage_XZ1', 'P027885675_CDS_0078___Vibrio_phage_vB_ValM_PVA8', 'P009931825_CDS_0263___Vibrio_phage_VH1_2019', 'P024381305_CDS_0078___Vibrio_phage_PVA8', 'P029686055_CDS_0384___Vibrio_phage_vB_VspM_VS2', 'VPD481_0304___Vibrio_phage_D481', 'VPD527_0304___Vibrio_phage_D527', 'P002757115_CDS_0421___Vibrio_phage_phi-Grn1', 'VPD490_0308___Vibrio_phage_D490', 'VP6E351A_0310___Vibrio_phage_6E35-1a', 'P012361135_CDS_0333___Vibrio_phage_V09', 'P000914495_CDS_0166___Vibrio_phage_VH7D', 'VPK491_0308___Vibrio_phage_K491', 'VPD484_0301___Vibrio_phage_D484', 'VPK566_0301___Vibrio_phage_K566', 'P014337555_CDS_0320___Vibrio_phage_vB_ValM_R11Z', 'VPD480_0304___Vibrio_phage_D480', 'P000910195_CDS_0132___Vibrio_phage_nt-1', 'VPD485_0306___Vibrio_phage_D485', 'VPK575_0307___Vibrio_phage_K575', 'P002757135_CDS_0142___Vibrio_phage_phi-ST2', 'P030927055_CDS_0078___Vibrio_phage_PC-Liy1', 'VPD525_0308___Vibrio_phage_D525', 'VPD486_0301___Vibrio_phage_D486', 'VPK571_0308___Vibrio_phage_K571', 'P014337545_CDS_0320___Vibrio_phage_vB_ValM_R10Z', 'P000843785_CDS_0263___Vibrio_phage_KVP40', 'P024234235_CDS_0093___Vibrio_phage_PVA23', 'VPF86_0222___Vibrio_phage_F86', 'P004771355_CDS_0278___Vibrio_phage_Va3', 'VPD491S_0304___Vibrio_phage_D491', 'P001501635_CDS_0416___Vibrio_phage_ValKK3', 'P020809685_CDS_0250___Vibrio_phage_VS1', 'VPK567_0303___Vibrio_phage_K567', 'VPD483_0308___Vibrio_phage_D483', 'VPD482_0310___Vibrio_phage_D482', 'P002630725_CDS_0258___Vibrio_phage_phi-pp2'}, 14: {'VPF86_0223___Vibrio_phage_F86'}, 15: {'VPF86_0224___Vibrio_phage_F86'}, 16: {'VPD480_0307___Vibrio_phage_D480', 'VPD482_0313___Vibrio_phage_D482', 'VPF86_0225___Vibrio_phage_F86', 'VPD481_0307___Vibrio_phage_D481', 'VPD527_0307___Vibrio_phage_D527', 'VPD486_0304___Vibrio_phage_D486', 'VPD484_0304___Vibrio_phage_D484', 'P029686055_CDS_0386___Vibrio_phage_vB_VspM_VS2', 'VPD485_0309___Vibrio_phage_D485', 'P009931825_CDS_0265___Vibrio_phage_VH1_2019', 'VPD525_0311___Vibrio_phage_D525', 'VPD490_0311___Vibrio_phage_D490', 'VPK571_0311___Vibrio_phage_K571', 'VPD483_0311___Vibrio_phage_D483', 'P002630725_CDS_0260___Vibrio_phage_phi-pp2', 'VPD491S_0307___Vibrio_phage_D491', 'P027885675_CDS_0081___Vibrio_phage_vB_ValM_PVA8', 'VPK491_0311___Vibrio_phage_K491', 'P000843785_CDS_0265___Vibrio_phage_KVP40', 'P012361135_CDS_0331___Vibrio_phage_V09', 'P020809685_CDS_0246___Vibrio_phage_VS1', 'P030927055_CDS_0081___Vibrio_phage_PC-Liy1', 'VP6E351A_0313___Vibrio_phage_6E35-1a', 'VPK575_0310___Vibrio_phage_K575', 'P024381305_CDS_0081___Vibrio_phage_PVA8'}, 17: {'VPD480_0308___Vibrio_phage_D480', 'P009931825_CDS_0266___Vibrio_phage_VH1_2019', 'VPK575_0311___Vibrio_phage_K575', 'VPD481_0308___Vibrio_phage_D481', 'P029686055_CDS_0387___Vibrio_phage_vB_VspM_VS2', 'P024381305_CDS_0082___Vibrio_phage_PVA8', 'P020809685_CDS_0245___Vibrio_phage_VS1', 'VPD484_0305___Vibrio_phage_D484', 'VPD525_0312___Vibrio_phage_D525', 'VPD490_0312___Vibrio_phage_D490', 'P000843785_CDS_0266___Vibrio_phage_KVP40', 'P030927055_CDS_0082___Vibrio_phage_PC-Liy1', 'VPD527_0308___Vibrio_phage_D527', 'VPK571_0312___Vibrio_phage_K571', 'P002630725_CDS_0261___Vibrio_phage_phi-pp2', 'P012361135_CDS_0330___Vibrio_phage_V09', 'VPD485_0310___Vibrio_phage_D485', 'VPD483_0312___Vibrio_phage_D483', 'VP6E351A_0314___Vibrio_phage_6E35-1a', 'VPF86_0226___Vibrio_phage_F86', 'VPK491_0312___Vibrio_phage_K491', 'VPD486_0305___Vibrio_phage_D486', 'P027885675_CDS_0082___Vibrio_phage_vB_ValM_PVA8', 'VPD491S_0308___Vibrio_phage_D491', 'VPD482_0314___Vibrio_phage_D482'}, 18: {'P012361135_CDS_0329___Vibrio_phage_V09', 'VPK571_0313___Vibrio_phage_K571', 'P009931825_CDS_0267___Vibrio_phage_VH1_2019', 'P029686055_CDS_0388___Vibrio_phage_vB_VspM_VS2', 'VPD527_0309___Vibrio_phage_D527', 'VPK491_0313___Vibrio_phage_K491', 'VPD484_0306___Vibrio_phage_D484', 'P000843785_CDS_0267___Vibrio_phage_KVP40', 'VPK575_0312___Vibrio_phage_K575', 'VPD485_0311___Vibrio_phage_D485', 'P002630725_CDS_0262___Vibrio_phage_phi-pp2', 'P024381305_CDS_0083___Vibrio_phage_PVA8', 'VPD491S_0309___Vibrio_phage_D491', 'VPD481_0309___Vibrio_phage_D481', 'VPD525_0313___Vibrio_phage_D525', 'VPD483_0313___Vibrio_phage_D483', 'P020809685_CDS_0244___Vibrio_phage_VS1', 'VPF86_0227___Vibrio_phage_F86', 'VPD480_0309___Vibrio_phage_D480', 'VPD482_0315___Vibrio_phage_D482', 'VPD490_0313___Vibrio_phage_D490', 'P030927055_CDS_0083___Vibrio_phage_PC-Liy1', 'VP6E351A_0315___Vibrio_phage_6E35-1a', 'P027885675_CDS_0083___Vibrio_phage_vB_ValM_PVA8', 'VPD486_0306___Vibrio_phage_D486'}, 19: {'P020809685_CDS_0243___Vibrio_phage_VS1', 'VPK491_0314___Vibrio_phage_K491', 'VPD525_0314___Vibrio_phage_D525', 'VPD486_0307___Vibrio_phage_D486', 'VPD485_0312___Vibrio_phage_D485', 'VPD481_0310___Vibrio_phage_D481', 'VPD490_0314___Vibrio_phage_D490', 'P012361135_CDS_0328___Vibrio_phage_V09', 'P024381305_CDS_0084___Vibrio_phage_PVA8', 'VPF86_0228___Vibrio_phage_F86', 'VP6E351A_0316___Vibrio_phage_6E35-1a', 'VPK575_0313___Vibrio_phage_K575', 'P000843785_CDS_0268___Vibrio_phage_KVP40', 'P009931825_CDS_0268___Vibrio_phage_VH1_2019', 'VPK571_0314___Vibrio_phage_K571', 'P030927055_CDS_0084___Vibrio_phage_PC-Liy1', 'P002630725_CDS_0263___Vibrio_phage_phi-pp2', 'VPD480_0310___Vibrio_phage_D480', 'VPD482_0316___Vibrio_phage_D482', 'VPD527_0310___Vibrio_phage_D527', 'VPD484_0307___Vibrio_phage_D484', 'VPD491S_0310___Vibrio_phage_D491', 'P027885675_CDS_0084___Vibrio_phage_vB_ValM_PVA8', 'P029686055_CDS_0389___Vibrio_phage_vB_VspM_VS2', 'VPD483_0314___Vibrio_phage_D483'}, 20: {'VPD490_0315___Vibrio_phage_D490', 'VPD485_0313___Vibrio_phage_D485', 'VPF86_0229___Vibrio_phage_F86', 'VPD486_0308___Vibrio_phage_D486', 'VPK571_0315___Vibrio_phage_K571', 'VPD491S_0311___Vibrio_phage_D491', 'VPD481_0311___Vibrio_phage_D481', 'VPK575_0314___Vibrio_phage_K575', 'VPD525_0315___Vibrio_phage_D525', 'VPD483_0315___Vibrio_phage_D483', 'VPK491_0315___Vibrio_phage_K491', 'VPD527_0311___Vibrio_phage_D527', 'VPD480_0311___Vibrio_phage_D480', 'VPD482_0317___Vibrio_phage_D482', 'VPD484_0308___Vibrio_phage_D484', 'VP6E351A_0317___Vibrio_phage_6E35-1a'}, 21: {'VPK491_0318___Vibrio_phage_K491', 'VPD490_0317___Vibrio_phage_D490', 'P000843785_CDS_0267___Vibrio_phage_KVP40', 'P012361135_CDS_0326___Vibrio_phage_V09', 'VPK491_0317___Vibrio_phage_K491', 'VPD485_0311___Vibrio_phage_D485', 'VPK575_0317___Vibrio_phage_K575', 'VPD486_0311___Vibrio_phage_D486', 'VPD525_0317___Vibrio_phage_D525', 'VPK567_0305___Vibrio_phage_K567', 'VPD483_0313___Vibrio_phage_D483', 'VPK566_0308___Vibrio_phage_K566', 'VPD490_0318___Vibrio_phage_D490', 'VPD480_0315___Vibrio_phage_D480', 'VPD482_0315___Vibrio_phage_D482', 'P020809685_CDS_0241___Vibrio_phage_VS1', 'VPF86_0232___Vibrio_phage_F86', 'P027885675_CDS_0083___Vibrio_phage_vB_ValM_PVA8', 'VPD486_0306___Vibrio_phage_D486', 'VPK567_0307___Vibrio_phage_K567', 'VP6E351A_0320___Vibrio_phage_6E35-1a', 'P012361135_CDS_0329___Vibrio_phage_V09', 'VPK567_0310___Vibrio_phage_K567', 'VPD491S_0313___Vibrio_phage_D491', 'VPK571_0313___Vibrio_phage_K571', 'VPK566_0303___Vibrio_phage_K566', 'VP6E351A_0319___Vibrio_phage_6E35-1a', 'VPF86_0231___Vibrio_phage_F86', 'VPD485_0315___Vibrio_phage_D485', 'VPD527_0313___Vibrio_phage_D527', 'VPK575_0312___Vibrio_phage_K575', 'VPD482_0320___Vibrio_phage_D482', 'P002630725_CDS_0262___Vibrio_phage_phi-pp2', 'P003927315_CDS_0276___Vibrio_phage_1.081.O._10N.286.52.C2', 'P027885675_CDS_0086___Vibrio_phage_vB_ValM_PVA8', 'P002630725_CDS_0267___Vibrio_phage_phi-pp2', 'P024381305_CDS_0083___Vibrio_phage_PVA8', 'VPD486_0310___Vibrio_phage_D486', 'VPD491S_0309___Vibrio_phage_D491', 'P030927055_CDS_0087___Vibrio_phage_PC-Liy1', 'VPD525_0313___Vibrio_phage_D525', 'P020809685_CDS_0244___Vibrio_phage_VS1', 'P009931825_CDS_0267___Vibrio_phage_VH1_2019', 'VPK575_0316___Vibrio_phage_K575', 'P009931825_CDS_0270___Vibrio_phage_VH1_2019', 'VPK566_0305___Vibrio_phage_K566', 'VPD491S_0314___Vibrio_phage_D491', 'P027885675_CDS_0087___Vibrio_phage_vB_ValM_PVA8', 'VPD484_0306___Vibrio_phage_D484', 'VPD525_0318___Vibrio_phage_D525', 'P012361135_CDS_0325___Vibrio_phage_V09', 'P029686055_CDS_0392___Vibrio_phage_vB_VspM_VS2', 'P024381305_CDS_0086___Vibrio_phage_PVA8', 'P000910195_CDS_0145___Vibrio_phage_nt-1', 'VPK566_0306___Vibrio_phage_K566', 'VPD483_0318___Vibrio_phage_D483', 'P000843785_CDS_0271___Vibrio_phage_KVP40', 'VPD480_0309___Vibrio_phage_D480', 'VPD527_0314___Vibrio_phage_D527', 'VPD483_0317___Vibrio_phage_D483', 'VPD481_0314___Vibrio_phage_D481', 'VPK571_0318___Vibrio_phage_K571', 'P030927055_CDS_0083___Vibrio_phage_PC-Liy1', 'P030927055_CDS_0086___Vibrio_phage_PC-Liy1', 'VP6E351A_0315___Vibrio_phage_6E35-1a', 'VPD481_0309___Vibrio_phage_D481', 'VPK567_0308___Vibrio_phage_K567', 'P002630725_CDS_0266___Vibrio_phage_phi-pp2', 'P029686055_CDS_0388___Vibrio_phage_vB_VspM_VS2', 'VPD527_0309___Vibrio_phage_D527', 'VPD480_0314___Vibrio_phage_D480', 'VPD482_0319___Vibrio_phage_D482', 'VPK571_0317___Vibrio_phage_K571', 'P020809685_CDS_0240___Vibrio_phage_VS1', 'VPK491_0313___Vibrio_phage_K491', 'VPD485_0316___Vibrio_phage_D485', 'P024381305_CDS_0087___Vibrio_phage_PVA8', 'VPD481_0313___Vibrio_phage_D481', 'VPD484_0310___Vibrio_phage_D484', 'P000843785_CDS_0270___Vibrio_phage_KVP40', 'P009931825_CDS_0271___Vibrio_phage_VH1_2019', 'P029686055_CDS_0391___Vibrio_phage_vB_VspM_VS2', 'VPD484_0311___Vibrio_phage_D484', 'VPD480_0313___Vibrio_phage_D480', 'VPF86_0227___Vibrio_phage_F86', 'VPD490_0313___Vibrio_phage_D490'}, 22: {'VPD525_0321___Vibrio_phage_D525', 'VPK575_0319___Vibrio_phage_K575', 'VPD491S_0316___Vibrio_phage_D491', 'VPD527_0316___Vibrio_phage_D527', 'VP6E351A_0322___Vibrio_phage_6E35-1a', 'VPD490_0320___Vibrio_phage_D490', 'VPF86_0233___Vibrio_phage_F86', 'VPK491_0320___Vibrio_phage_K491', 'VPD481_0316___Vibrio_phage_D481', 'VPD482_0322___Vibrio_phage_D482', 'VPD480_0317___Vibrio_phage_D480', 'VPD484_0314___Vibrio_phage_D484', 'VPD485_0318___Vibrio_phage_D485', 'VPK571_0320___Vibrio_phage_K571', 'VPD486_0313___Vibrio_phage_D486', 'VPD483_0320___Vibrio_phage_D483'}, 23: {'VPF86_0234___Vibrio_phage_F86', 'VPK571_0321___Vibrio_phage_K571', 'VPD525_0322___Vibrio_phage_D525', 'VPK491_0321___Vibrio_phage_K491', 'VPD483_0321___Vibrio_phage_D483', 'VPD491S_0317___Vibrio_phage_D491', 'VPK575_0320___Vibrio_phage_K575', 'VPD480_0318___Vibrio_phage_D480', 'VPD484_0315___Vibrio_phage_D484', 'VPD490_0321___Vibrio_phage_D490', 'VPD527_0317___Vibrio_phage_D527', 'VPD482_0323___Vibrio_phage_D482', 'VP6E351A_0323___Vibrio_phage_6E35-1a', 'VPD481_0317___Vibrio_phage_D481', 'VPD485_0319___Vibrio_phage_D485', 'VPD486_0314___Vibrio_phage_D486'}, 24: {'P000843785_CDS_0272___Vibrio_phage_KVP40', 'VPK491_0322___Vibrio_phage_K491', 'VPD482_0324___Vibrio_phage_D482', 'VPD527_0318___Vibrio_phage_D527', 'P012361135_CDS_0324___Vibrio_phage_V09', 'VPK571_0322___Vibrio_phage_K571', 'VPD490_0322___Vibrio_phage_D490', 'VPD491S_0318___Vibrio_phage_D491', 'P027885675_CDS_0088___Vibrio_phage_vB_ValM_PVA8', 'VPD480_0319___Vibrio_phage_D480', 'VPD484_0316___Vibrio_phage_D484', 'VPD481_0318___Vibrio_phage_D481', 'VPD525_0323___Vibrio_phage_D525', 'VPD485_0320___Vibrio_phage_D485', 'VPD486_0315___Vibrio_phage_D486', 'P030927055_CDS_0088___Vibrio_phage_PC-Liy1', 'P009931825_CDS_0272___Vibrio_phage_VH1_2019', 'P020809685_CDS_0239___Vibrio_phage_VS1', 'VPK575_0321___Vibrio_phage_K575', 'VP6E351A_0324___Vibrio_phage_6E35-1a', 'VPD483_0322___Vibrio_phage_D483', 'VPF86_0235___Vibrio_phage_F86', 'P024381305_CDS_0088___Vibrio_phage_PVA8', 'P002630725_CDS_0268___Vibrio_phage_phi-pp2', 'P029686055_CDS_0393___Vibrio_phage_vB_VspM_VS2'}, 25: {'VPF86_0236___Vibrio_phage_F86'}, 26: {'VPD485_0322___Vibrio_phage_D485', 'P012361135_CDS_0322___Vibrio_phage_V09', 'VPD491S_0320___Vibrio_phage_D491', 'VPK491_0324___Vibrio_phage_K491', 'P024381305_CDS_0090___Vibrio_phage_PVA8', 'VPD525_0325___Vibrio_phage_D525', 'VPD484_0318___Vibrio_phage_D484', 'VPK571_0324___Vibrio_phage_K571', 'VPD481_0320___Vibrio_phage_D481', 'P027885675_CDS_0090___Vibrio_phage_vB_ValM_PVA8', 'VPD486_0317___Vibrio_phage_D486', 'VPD483_0324___Vibrio_phage_D483', 'VPK575_0323___Vibrio_phage_K575', 'VP6E351A_0326___Vibrio_phage_6E35-1a', 'P030927055_CDS_0090___Vibrio_phage_PC-Liy1', 'VPD480_0321___Vibrio_phage_D480', 'VPD490_0324___Vibrio_phage_D490', 'P020809685_CDS_0237___Vibrio_phage_VS1', 'P000843785_CDS_0274___Vibrio_phage_KVP40', 'P009931825_CDS_0274___Vibrio_phage_VH1_2019', 'VPD482_0326___Vibrio_phage_D482', 'VPF86_0237___Vibrio_phage_F86', 'VPD527_0320___Vibrio_phage_D527', 'P002630725_CDS_0269___Vibrio_phage_phi-pp2', 'P029686055_CDS_0394___Vibrio_phage_vB_VspM_VS2'}, 27: {'VPK566_0302___Vibrio_phage_K566', 'VPK567_0304___Vibrio_phage_K567'}, 28: {'VPK566_0304___Vibrio_phage_K566', 'VPK567_0306___Vibrio_phage_K567'}, 29: {'VPK567_0309___Vibrio_phage_K567', 'P000910195_CDS_0147___Vibrio_phage_nt-1', 'VPK566_0307___Vibrio_phage_K566'}, 30: {'VPK567_0311___Vibrio_phage_K567', 'VPK566_0309___Vibrio_phage_K566'}, 31: {'VPK566_0310___Vibrio_phage_K566', 'VPK567_0312___Vibrio_phage_K567'}, 32: {'VPK567_0313___Vibrio_phage_K567', 'VPK566_0311___Vibrio_phage_K566'}, 33: {'VPK566_0312___Vibrio_phage_K566', 'VPK567_0314___Vibrio_phage_K567'}, 34: {'VPK567_0315___Vibrio_phage_K567', 'VPK566_0313___Vibrio_phage_K566'}, 35: {'VPK567_0316___Vibrio_phage_K567', 'VPK566_0314___Vibrio_phage_K566'}, 36: {'VPD480_0305___Vibrio_phage_D480', 'VPK571_0309___Vibrio_phage_K571', 'VPD481_0305___Vibrio_phage_D481', 'VPD484_0302___Vibrio_phage_D484', 'VPD486_0302___Vibrio_phage_D486', 'VPD490_0309___Vibrio_phage_D490', 'VPD527_0305___Vibrio_phage_D527', 'VPD491S_0305___Vibrio_phage_D491', 'VPD485_0307___Vibrio_phage_D485', 'VPD482_0311___Vibrio_phage_D482', 'VPD525_0309___Vibrio_phage_D525', 'VP6E351A_0311___Vibrio_phage_6E35-1a', 'VPK575_0308___Vibrio_phage_K575', 'VPK491_0309___Vibrio_phage_K491', 'VPD483_0309___Vibrio_phage_D483'}, 37: {'VPD491S_0306___Vibrio_phage_D491', 'VPD484_0303___Vibrio_phage_D484', 'VPD482_0312___Vibrio_phage_D482', 'VPD486_0303___Vibrio_phage_D486', 'VPD485_0308___Vibrio_phage_D485', 'VPD490_0310___Vibrio_phage_D490', 'VPD481_0306___Vibrio_phage_D481', 'VPD480_0306___Vibrio_phage_D480', 'VPD527_0306___Vibrio_phage_D527'}, 38: {'VPK575_0318___Vibrio_phage_K575', 'VPD490_0319___Vibrio_phage_D490', 'VPD484_0313___Vibrio_phage_D484', 'VPD491S_0315___Vibrio_phage_D491', 'VPD482_0321___Vibrio_phage_D482', 'VPK571_0319___Vibrio_phage_K571', 'VPD525_0320___Vibrio_phage_D525', 'VPD483_0319___Vibrio_phage_D483', 'VPD481_0315___Vibrio_phage_D481', 'VPD486_0312___Vibrio_phage_D486', 'VPD480_0316___Vibrio_phage_D480', 'VPD485_0317___Vibrio_phage_D485', 'VPD527_0315___Vibrio_phage_D527', 'VPK491_0319___Vibrio_phage_K491', 'VP6E351A_0321___Vibrio_phage_6E35-1a'}, 39: {'VPD484_0317___Vibrio_phage_D484', 'VPD486_0316___Vibrio_phage_D486', 'VPK491_0323___Vibrio_phage_K491', 'VPD483_0323___Vibrio_phage_D483', 'VPD525_0324___Vibrio_phage_D525', 'VPD527_0319___Vibrio_phage_D527', 'VPK571_0323___Vibrio_phage_K571', 'VPD491S_0319___Vibrio_phage_D491', 'VPD485_0321___Vibrio_phage_D485', 'VPD490_0323___Vibrio_phage_D490', 'VPD480_0320___Vibrio_phage_D480', 'VPD482_0325___Vibrio_phage_D482', 'VPK575_0322___Vibrio_phage_K575', 'VPD481_0319___Vibrio_phage_D481', 'VP6E351A_0325___Vibrio_phage_6E35-1a'}, 40: {'VP6E351A_0327___Vibrio_phage_6E35-1a', 'VPD483_0325___Vibrio_phage_D483', 'P020809685_CDS_0236___Vibrio_phage_VS1', 'P002630725_CDS_0270___Vibrio_phage_phi-pp2', 'VPD525_0326___Vibrio_phage_D525', 'VPK575_0324___Vibrio_phage_K575', 'VPK571_0325___Vibrio_phage_K571', 'VPK491_0325___Vibrio_phage_K491', 'VPD485_0323___Vibrio_phage_D485', 'P029686055_CDS_0395___Vibrio_phage_vB_VspM_VS2'}, 41: {'VPK575_0309___Vibrio_phage_K575', 'VPD483_0310___Vibrio_phage_D483', 'VP6E351A_0312___Vibrio_phage_6E35-1a', 'VPD525_0310___Vibrio_phage_D525', 'VPK571_0310___Vibrio_phage_K571', 'VPK491_0310___Vibrio_phage_K491'}, 42: {'VPD525_0319___Vibrio_phage_D525', 'VPD484_0312___Vibrio_phage_D484'}, 43: {'P000910195_CDS_0130___Vibrio_phage_nt-1'}, 44: {'P000910195_CDS_0131___Vibrio_phage_nt-1'}, 45: {'P000910195_CDS_0135___Vibrio_phage_nt-1'}, 46: {'P000910195_CDS_0136___Vibrio_phage_nt-1'}, 47: {'P000910195_CDS_0137___Vibrio_phage_nt-1'}, 48: {'P000910195_CDS_0138___Vibrio_phage_nt-1'}, 49: {'P000910195_CDS_0139___Vibrio_phage_nt-1'}, 50: {'P000910195_CDS_0140___Vibrio_phage_nt-1'}, 51: {'P000910195_CDS_0142___Vibrio_phage_nt-1'}, 52: {'P000910195_CDS_0143___Vibrio_phage_nt-1'}, 53: {'P000910195_CDS_0144___Vibrio_phage_nt-1'}, 54: {'P000910195_CDS_0146___Vibrio_phage_nt-1'}, 55: {'P030927055_CDS_0079___Vibrio_phage_PC-Liy1', 'P009931825_CDS_0264___Vibrio_phage_VH1_2019', 'P000843785_CDS_0264___Vibrio_phage_KVP40', 'P012361135_CDS_0332___Vibrio_phage_V09', 'P024381305_CDS_0079___Vibrio_phage_PVA8', 'P020809685_CDS_0249___Vibrio_phage_VS1', 'P002630725_CDS_0259___Vibrio_phage_phi-pp2', 'P027885675_CDS_0079___Vibrio_phage_vB_ValM_PVA8'}, 56: {'P020809685_CDS_0248___Vibrio_phage_VS1'}, 57: {'P020809685_CDS_0247___Vibrio_phage_VS1'}, 58: {'P020809685_CDS_0238___Vibrio_phage_VS1'}, 59: {'P029686055_CDS_0385___Vibrio_phage_vB_VspM_VS2', 'P009931825_CDS_0264___Vibrio_phage_VH1_2019', 'P000843785_CDS_0264___Vibrio_phage_KVP40', 'P012361135_CDS_0332___Vibrio_phage_V09', 'P002630725_CDS_0259___Vibrio_phage_phi-pp2'}, 60: {'P012361135_CDS_0323___Vibrio_phage_V09', 'P024381305_CDS_0089___Vibrio_phage_PVA8', 'P009931825_CDS_0273___Vibrio_phage_VH1_2019', 'P027885675_CDS_0089___Vibrio_phage_vB_ValM_PVA8', 'P000843785_CDS_0273___Vibrio_phage_KVP40', 'P030927055_CDS_0089___Vibrio_phage_PC-Liy1'}, 61: {'P012361135_CDS_0321___Vibrio_phage_V09', 'P027885675_CDS_0091___Vibrio_phage_vB_ValM_PVA8', 'P009931825_CDS_0275___Vibrio_phage_VH1_2019', 'P030927055_CDS_0091___Vibrio_phage_PC-Liy1', 'P000843785_CDS_0275___Vibrio_phage_KVP40', 'P024381305_CDS_0091___Vibrio_phage_PVA8'}, 62: {'P024381305_CDS_0080___Vibrio_phage_PVA8', 'P027885675_CDS_0080___Vibrio_phage_vB_ValM_PVA8', 'P030927055_CDS_0080___Vibrio_phage_PC-Liy1'}, 63: {'P000914495_CDS_0168___Vibrio_phage_VH7D', 'P002757115_CDS_0423___Vibrio_phage_phi-Grn1', 'P002757135_CDS_0144___Vibrio_phage_phi-ST2', 'P022984185_CDS_0149___Vibrio_phage_XZ1', 'P001501635_CDS_0414___Vibrio_phage_ValKK3', 'P004771355_CDS_0276___Vibrio_phage_Va3', 'P014337545_CDS_0318___Vibrio_phage_vB_ValM_R10Z', 'P014337555_CDS_0318___Vibrio_phage_vB_ValM_R11Z', 'P024234235_CDS_0095___Vibrio_phage_PVA23'}}
    dicoMyColor = {}
    for num in dicoBlast:
        color = random_hex_color()
        for elem in dicoBlast[num]:
            dicoMyColor[elem.split("___")[0]] = color
    lstFiles, maxpathSize = get_input_files(pathIN, "rbh_linear_plot", [ext])
    if len(lstFiles) == 0:
        printcolor("[ERROR: rbh_linear_plot]\nAny input files found\n", 1, "212;64;89", "None", True)
        exit_gemini()
    pathCLUSTER = path_converter(pathCLUSTER)
    pathOUT = path_converter(pathOUT)
    if pathSUBCORE != "None":
        dicoSubCore = {}
        pathSUBCORE = path_converter(pathSUBCORE)
        for line in read_file(pathSUBCORE):
            try:
                dicoSubCore[line.split("\t")[0]]['set'].add(line.split("\t")[1])
                dicoSubCore[line.split("\t")[0]]['color'] = line.split("\t")[2]
            except KeyError:
                dicoSubCore[line.split("\t")[0]] = {'set':set([line.split("\t")[1]]), 'color':line.split("\t")[2]}
    os.makedirs(pathOUT, exist_ok=True)
    os.makedirs(pathOUT+"/PNG", exist_ok=True)
    os.makedirs(pathOUT+"/SVG", exist_ok=True)
    # Make dicoLTcolor based on RBH clusters
    printcolor("♊ Cluster colors"+"\n")
    dicoLTcolor = {}
    dicoCLUSTER = load_json(pathCLUSTER)
    dicoProtToCluster = {}
    dicoClusterToOrg = {}
    dicoClusterColor = {}
    dicoClusterType = {}
    setColor = set()
    setAvailableOrg = set()
    # If distinct color for each gene family use random_hex_color
    # Else simply use a gradient rare to frequent gene
    if distinctColor is False:
        HEX_gradient, RBG_gradient = linear_gradient(start_hex="#e6e6e6", finish_hex="#333333", n=len(lstFiles))
    for cluster in dicoCLUSTER:
        dicoClusterToOrg[cluster] = {}
        setOrg = set()  # To avoid paralogous
        for header in dicoCLUSTER[cluster]:
            lt = header.split(" [")[0].split("|")[0].split(" ")[0]
            org = header.split(" [")[1].replace("]", "")
            setOrg.add(org)
            setAvailableOrg.add(org)
            dicoClusterToOrg[cluster][org] = lt
            try:
                dicoProtToCluster[org][lt] = cluster
            except KeyError:
                dicoProtToCluster[org] = {lt: cluster}
        # Define cluster color and type
        if distinctColor is True:
            # Singleton gene > white
            if len(setOrg) == 1:
                color = "#ffffff"
                clusterType = "singleton"
            # Core gene > dark red
            elif len(setOrg) == len(lstFiles):
                color = "#aa0000"
                clusterType = "core"
            else:
                color = random_hex_color()
                clusterType = cluster
        else:
            # Sub-core gene
            foundSubcore = False
            if pathSUBCORE != "None":
                for group in dicoSubCore:
                    # if set(setOrg) == dicoSubCore[group]['set']:
                    if len(setOrg-dicoSubCore[group]['set']) == 0 and len(setOrg)>=len(dicoSubCore[group]['set'])*0.9:
                        color = dicoSubCore[group]['color']
                        clusterType = "subcore"
                        foundSubcore = True
            if foundSubcore is False:
                # Singleton gene > black
                if len(setOrg) == 1:
                    color = "#000000"
                    clusterType = "singleton"
                # Core gene > white
                # elif len(setOrg) == len(lstFiles):
                elif len(setOrg) >= len(lstFiles)*0.9:
                    color = "#ffffff"
                    clusterType = "core"
                # Accessory gene > grey
                else:
                    color = "#b3b3b3"
                    clusterType = "accessory ("+str(len(setOrg))+")"
        dicoClusterColor[cluster] = color
        dicoClusterType[cluster] = clusterType
        setColor.add(color)
        # Apply to all LT
        for header in dicoCLUSTER[cluster]:
            lt = header.split(" [")[0].split("|")[0].split(" ")[0]
            dicoLTcolor[lt] = color
    # Search longest genome
    dicoAllGFF = {}
    printcolor("♊ Max length"+"\n")
    maxSeqLen = 0
    for pathGFF in lstFiles:
        file = os.path.basename(pathGFF)
        orgName = file.replace(ext, "").replace("."+ext, "")
        # ***** PARSE GFF ***** #
        dicoAllGFF[orgName] = make_gff_dict(pathIN=pathGFF, ext=ext)[orgName]
        maxSeqLen = max(maxSeqLen, dicoAllGFF[orgName]['length'])
    maxSeqLen = 17500
    # Browse GFF3
    printcolor("♊ Plotting"+"\n")
    pbar = tqdm(total=int(len(lstFiles)), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    setSVGfiles = set()

    for orgName in dicoAllGFF:
        pathPNG = pathOUT+"/PNG/"+orgName+".png"
        pathSVG = pathOUT+"/SVG/"+orgName+".svg"
        setSVGfiles.add(pathSVG)
        pbar.set_description_str(orgName+" ".rjust(maxpathSize-len(orgName)))
        features = []
        # ***** BROWSE GENES ***** #
        for geneType in dicoAllGFF[orgName]:
            if geneType not in ['length','pseudogene']:
                for geneEntry in dicoAllGFF[orgName][geneType]:
                    if geneType == "tRNA" or "tRNA" in geneEntry['attributes']['locus_tag']:  # green
                        color = "#2ca05a"
                    elif 'locus_tag' in geneEntry['attributes'] and geneEntry['attributes']['locus_tag'].replace(" ", "") in dicoLTcolor:
                        color = dicoLTcolor[geneEntry['attributes']['locus_tag'].replace(" ", "")]
                    elif 'protein_id' in geneEntry['attributes'] and geneEntry['attributes']['protein_id'].replace(" ", "") in dicoLTcolor:
                        color = dicoLTcolor[geneEntry['attributes']['protein_id'].replace(" ", "")]
                    else:
                        color = "#0000ff" # missing singleton
                    color = dicoMyColor[geneEntry['attributes']['locus_tag'].replace(" ", "")]
                    geneFeature = GraphicFeature(start=geneEntry['start'], end=geneEntry['end'], strand=int(geneEntry['strand']+"1"), color=color, linewidth=0)
                    features.append(geneFeature)
        # ***** PLOT GENES ***** #
        record = GraphicRecord(sequence_length=maxSeqLen, features=features, first_index=-100)
        ax, _ = record.plot(figure_width=50, )
        ax.figure.savefig(pathPNG, dpi=300)
        ax.figure.savefig(pathSVG)
        plt.close('all')
        pbar.update(1)
        title("Plotting", pbar)
    pbar.close()
    # ***** Reformat SVG output  ***** # (selectable color / Delete blank, line and axis / valign)
    printcolor("♊ Reformat SVG"+"\n")
    pbar = tqdm(total=len(setSVGfiles), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    for pathSVG in setSVGfiles:
        pbar.set_description_str(os.path.basename(pathSVG)+" ".rjust(maxpathSize-len(os.path.basename(pathSVG))))
        # Read initial SVG
        SVG = open(pathSVG, 'r')
        dataSVG = SVG.read()
        splitG = dataSVG.split("<g")
        SVG.close()
        dicoPatchToColor = {}
        for partG in splitG:
            if "id=\"patch_" in partG:
                splitLine = partG.split("\n")
                patchID = ""
                color = ""
                for line in splitLine:
                    if "id=\"patch_" in line:
                        patchID = line.split("\"")[1]
                    if "style=\"fill:" in line:
                        color = line.split(": ")[1].split(";")[0].split("\"")[0]
                if color != "" and color[0] == "#":
                    dicoPatchToColor[patchID] = color
        # Apply modifications
        SVG = open(pathSVG, 'w')
        for patchID in dicoPatchToColor:
            dataSVG = dataSVG.replace("<g id=\""+patchID+"\">", "<g id=\""+patchID+"\"\nstyle=\"fill:"+dicoPatchToColor[patchID]+"\">")
            dataSVG = dataSVG.replace(dicoPatchToColor[patchID], dicoPatchToColor[patchID]+";fill-opacity:1")
        SVG.write(dataSVG)
        SVG.close()


# <g id="patch_3">
#     <path d="M 1888.100538 130.552 
# Q 1845.912206 130.552 1803.723874 130.552 
# L 1803.723874 130.552 
# Q 1799.523586 127.052 1795.323298 123.552 
# Q 1799.523586 120.052 1803.723874 116.552 
# L 1803.723874 116.552 
# Q 1845.912206 116.552 1888.100538 116.552 
# L 1888.100538 130.552 
# z
# " clip-path="url(#p958b9d9612)" style="fill: #3c3c3c"/>
#    </g>


        # Delete blank, line and axis
        try:
            tree = ET.parse(pathSVG)
        except: print("tree = ET.parse("+pathSVG+")") ; exit_gemini
        root = tree.getroot()
        lst_block_id = ["patch_1", "line2d_1", "matplotlib.axis_1"]
        for block_id in lst_block_id:
            element_to_remove = root.find(".//*[@id='" + block_id + "']")
            if element_to_remove is not None:
                parent_element = None
                for elem in root.iter():
                    if element_to_remove in elem:
                        parent_element = elem
                        break
                if parent_element is not None:
                    parent_element.remove(element_to_remove)
        tree.write(pathSVG)
        # Align all genes
        xvfb = Xvfb()
        with xvfb:  # Start the virtual framebuffer for inkscape gui dependent verb
            os.system("inkscape --verb=EditSelectAll --verb=SelectionUnGroup --verb=EditSelectAll --verb=SelectionUnGroup \
                       --verb=AlignVerticalCenter --verb=EditSelectAll --verb=SelectionGroup \
                       --verb=FileSave --verb=FileQuit "+pathSVG)
        pbar.update(1)
        title("Reformat", pbar)
    pbar.close()
    # ***** Ordering clusters  ***** #
    printcolor("♊ Ordering clusters"+"\n")
    dicoOrder = OrderedDict()
    # Initialize dicoOrder with a first organism cluster list
    for lt in sorted(dicoProtToCluster[list(setAvailableOrg)[0]].keys()):
        dicoOrder[len(dicoOrder)] = dicoProtToCluster[list(setAvailableOrg)[0]][lt]
    # Browse others organism cluster list
    for orgName in list(setAvailableOrg)[1:]:
        lastPos = 0
        for lt in sorted(dicoProtToCluster[orgName].keys()):
            cluster = dicoProtToCluster[orgName][lt]
            lstOrderedClustered = list(dicoOrder.values())
            found = False
            for i in range(len(lstOrderedClustered)):
                if lstOrderedClustered[i] == cluster:
                    found = True
                    break
            if found is True:
                lastPos = i
            else:
                lastPos += 1
                dicoOrder = shift_dico_pos(dicoOrder, cluster, lastPos)
    # ***** Ordering clusters  ***** # (required same orientation and same first gene)
    printcolor("♊ Ordered clusters table"+"\n")
    dicoWidth = {0: 0, 1: 0}
    workbook = xlsxwriter.Workbook(pathOUT+"/ordered_clusters.xlsx")
    worksheet = workbook.add_worksheet()
    # Header and row format (row bgcolor correspond to plot colors)
    headerFormat = workbook.add_format({'align': 'center', 'valign': 'bottom', 'border': 1, 'font_size': 11, 'bold': True})
    headerFormatOrg = workbook.add_format({'align': 'center', 'valign': 'bottom', 'border': 1, 'font_size': 11, 'bold': True})
    headerFormatOrg.set_rotation(90)
    dicoRowFormat = {}
    dicoHeaderClusterFormat = {}
    for bgColor in setColor:
        dicoHeaderClusterFormat[bgColor] = workbook.add_format({'align': 'center', 'valign': 'bottom', 'border': 1, 'font_size': 11, 'bold': True, 'bg_color': bgColor})
        dicoRowFormat[bgColor] = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 8, 'bg_color': bgColor})
        if requires_white_text(bgColor):
            dicoHeaderClusterFormat[bgColor].set_font_color('#FFFFFF')
            dicoRowFormat[bgColor].set_font_color('#FFFFFF')
        else:
            dicoHeaderClusterFormat[bgColor].set_font_color('#000000')
            dicoRowFormat[bgColor].set_font_color('#000000')
    # Write header
    worksheet.write(0, 0, "Cluster", headerFormat)
    worksheet.write(0, 1, "Type", headerFormat)
    col = 2
    for orgName in setAvailableOrg:
        worksheet.write(0, col, orgName, headerFormatOrg)
        col += 1
    row = 1
    # Write row
    for order in dicoOrder:
        cluster = dicoOrder[order]
        rowColor = dicoClusterColor[cluster]
        headerFormat = dicoHeaderClusterFormat[rowColor]
        rowFormat = dicoRowFormat[rowColor]
        worksheet.write(row, 0, cluster, headerFormat)
        dicoWidth[0] = max(dicoWidth[0], len(cluster))
        worksheet.write(row, 1, dicoClusterType[cluster], headerFormat)
        dicoWidth[1] = max(dicoWidth[1], len(dicoClusterType[cluster]))
        col = 2
        for orgName in setAvailableOrg:
            if orgName in dicoClusterToOrg[cluster]:
                cellValue = dicoClusterToOrg[cluster][orgName]
            else:
                cellValue = "Nan"
            worksheet.write(row, col, cellValue, rowFormat)
            try:
                dicoWidth[col] = max(dicoWidth[col], len(str(cellValue)))
            except KeyError:
                dicoWidth[col] = len(str(cellValue))
            col += 1
        row += 1
    # Adjust row height and column width
    for col in dicoWidth:
        worksheet.set_column(col, col, dicoWidth[col])
    workbook.close()


@fct_checker
def gff_to_linear_group_geneplot(pathIN: str, pathCLUSTER: str, pathGROUP: str, pathOUT: str, ext: str = ".gff") -> Tuple[str, str, str, str, str]:
    '''
     ------------------------------------------------------------
    |             GFF3 TO LINEAR GENE PLOT PER GROUP             |
    |------------------------------------------------------------|
    | Create linear gene plot from GFF3 file and merge per group |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN     : path of input files or folder (required)   |
    |    pathOUT    : path of output files (required)            |
    |    pathCLUSTER: path of JSON rbh cluster (required)        |
    |    pathGROUP  : path of input tab group file (required)    |
    |    ext        : extension of input files (default=.gff)    |
     ------------------------------------------------------------
     input tab group: <name>\t<group>\t<subgroup color>
     singleton => black
     core => grey
    '''
    lstFiles, maxpathSize = get_input_files(pathIN, "gff_to_linear_group_geneplot", [ext])
    if len(lstFiles) == 0:
        printcolor("[ERROR: gff_to_linear_group_geneplot]\nAny input files found\n", 1, "212;64;89", "None", True)
        exit_gemini()
    pathCLUSTER = path_converter(pathCLUSTER)
    pathOUT = path_converter(pathOUT)
    pathGROUP = path_converter(pathGROUP)
    os.makedirs(pathOUT, exist_ok=True)
    # ***** PARSE tabular group file ***** #
    printcolor("♊ Read groups"+"\n")
    dicoOrgToGroup = {}
    dicoGroupToOrg = {}
    dicoOrgToSubGroupColor = {}
    GROUP = open(pathGROUP, 'r')
    lstLines = GROUP.read().split("\n")
    GROUP.close()
    for line in lstLines:
        if line != "":
            orgName = line.split("\t")[0]
            groupName = line.split("\t")[1]
            subGroupColor = line.split("\t")[2]
            dicoOrgToGroup[orgName] = groupName
            dicoOrgToSubGroupColor[orgName] = subGroupColor
            try:
                dicoGroupToOrg[groupName].append(orgName)
            except KeyError:
                dicoGroupToOrg[groupName] = [orgName]
    # ***** CLUSTERING genes ***** #
    printcolor("♊ Read clustering"+"\n")
    dicoCLUSTER = load_json(pathCLUSTER)
    printcolor("♊ Group clustering"+"\n")
    dicoGFF = make_gff_dict(pathIN=pathIN, ext=ext)
    dicoLTtoOrg = {}
    dicoGroupToOrgLT = {}
    dicoLTcolor = {}
    seqLen = 0
    for orgName in dicoOrgToGroup:
        for geneType in dicoGFF[orgName]:
            if geneType == 'length':
                seqLen = max(seqLen, dicoGFF[orgName][geneType])
            if geneType != 'length':
                for geneEntry in dicoGFF[orgName][geneType]:
                    if not geneType[0] == "#":
                        lt = geneEntry['attributes']['locus_tag']
                        dicoGroupToOrgLT[lt] = dicoOrgToGroup[orgName]
                        dicoLTtoOrg[lt] = orgName
    for group in dicoGroupToOrg:
        for cluster in dicoCLUSTER:
            setOrg = set()  # To avoid paralogous
            setOrgSubGroupColor = set()
            for header in dicoCLUSTER[cluster]:
                lt = header.split(" [")[0].split("|")[0]
                if lt in dicoGroupToOrgLT and dicoGroupToOrgLT[lt] == group:
                    setOrg.add(dicoLTtoOrg[lt])
                    setOrgSubGroupColor.add(dicoOrgToSubGroupColor[dicoLTtoOrg[lt]])
            # Singleton gene in black
            if len(setOrg) == 1:
                color = "#1a1a1a"
            # If group core gene => grey
            elif len(setOrg) == len(dicoGroupToOrg[group]):
                color = "#cccccc"
            # Look for core subgroup
            elif len(setOrgSubGroupColor) == 1 and list(setOrgSubGroupColor)[0].lower() != "none":
                color = list(setOrgSubGroupColor)[0]
            # Else random color
            else:
                color = random_hex_color(False)
            # Apply to all LT
            for header in dicoCLUSTER[cluster]:
                lt = header.split(" [")[0].split("|")[0]
                dicoLTcolor[lt] = color
    # ***** BROWSE input GFF files ***** #
    dicoGroupToOrgPlot = {}
    printcolor("♊ GraphicRecords"+"\n")
    pbar = tqdm(total=int(len(lstFiles)), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    for pathGFF in lstFiles:
        file = os.path.basename(pathGFF)
        orgName = file.replace(ext, "").replace("."+ext, "")
        if orgName in dicoOrgToGroup:
            pbar.set_description_str(orgName+" ".rjust(maxpathSize-len(orgName)))
            features = []
            # ***** PARSE GFF ***** #
            dicoGFF = make_gff_dict(pathIN=pathGFF, ext=ext)
            # ***** BROWSE GENES ***** #
            for geneType in dicoGFF[orgName]:
                if geneType != 'length':
                    for geneEntry in dicoGFF[orgName][geneType]:
                        if geneType in ["CDS", "tRNA"]:
                            color = dicoLTcolor[geneEntry['attributes']['locus_tag']]
                        else:
                            continue
                        geneFeature = GraphicFeature(start=geneEntry['start'], end=geneEntry['end'],
                                                     strand=int(geneEntry['strand']+"1"), color=color, linewidth=0)
                        features.append(geneFeature)
            # ***** PLOT GENES ***** #
            record = GraphicRecord(sequence_length=seqLen, features=features)
            try:
                dicoGroupToOrgPlot[dicoOrgToGroup[orgName]].append((orgName, record))
            except KeyError:
                dicoGroupToOrgPlot[dicoOrgToGroup[orgName]] = [(orgName, record)]
        pbar.update(1)
        title("GraphicRecords", pbar)
    pbar.close()
    # ***** Unify group plot ***** #
    printcolor("♊ Unify plots"+"\n")
    pbar = tqdm(total=int(len(dicoGroupToOrgPlot)), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    dicoGroupPNG = {}
    setSVGfiles = set()
    for group in dicoGroupToOrg:
        if len(dicoGroupToOrg[group]) >= 2:
            pathOUTgroup = pathOUT+"/"+group
            os.makedirs(pathOUTgroup, exist_ok=True)
            pbar.set_description_str("Group = "+group+" ".rjust(maxpathSize-len("Group = "+group)))
            dicoGroupPNG[group] = []
            # Get longest sequence for current group and enlarge to 10%
            maxLen = 0
            for (orgName, record) in dicoGroupToOrgPlot[group]:
                maxLen = max(record.sequence_length, maxLen)
            maxLen = maxLen+int(maxLen/10)
            # Get GraphicFeature
            for (orgName, record) in dicoGroupToOrgPlot[group]:
                pathPNG = pathOUTgroup+"/"+orgName+".png"
                pathSVG = pathOUTgroup+"/"+orgName+".svg"
                ax, _ = record.plot(figure_width=50)
                ax.figure.savefig(pathPNG, dpi=300)
                ax.figure.savefig(pathSVG, backend="SVG")
                setSVGfiles.add(pathSVG)
                plt.close('all')
        pbar.update(1)
        title("Unify", pbar)
    pbar.close()
    # ***** Reformat SVG output  ***** # (selectable color)
    printcolor("♊ Reformat SVG"+"\n")
    pbar = tqdm(total=len(setSVGfiles), dynamic_ncols=True, ncols=50+maxpathSize, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt} [{desc}]")
    for pathSVG in setSVGfiles:
        pbar.set_description_str(os.path.basename(pathSVG)+" ".rjust(maxpathSize-len(os.path.basename(pathSVG))))
        # Read initial SVG
        SVG = open(pathSVG, 'r')
        dataSVG = SVG.read()
        splitG = dataSVG.split("<g")
        SVG.close()
        dicoPatchToColor = {}
        for partG in splitG:
            if "id=\"patch_" in partG:
                splitLine = partG.split("\n")
                patchID = ""
                color = ""
                for line in splitLine:
                    if "id=\"patch_" in line:
                        patchID = line.split("\"")[1]
                    if "style=\"fill:" in line:
                        color = line.split(":")[1].split(";")[0].replace("\"", "")
                dicoPatchToColor[patchID] = color
        # Apply modifications
        SVG = open(pathSVG, 'w')
        for patchID in dicoPatchToColor:
            dataSVG = dataSVG.replace("<g id=\""+patchID+"\">", "<g id=\""+patchID+"\"\nstyle=\"fill:"+dicoPatchToColor[patchID]+"\">")
            dataSVG = dataSVG.replace(dicoPatchToColor[patchID], dicoPatchToColor[patchID]+";fill-opacity:1")
        SVG.write(dataSVG)
        SVG.close()
        pbar.update(1)
        title("Reformat", pbar)
    pbar.close()


@fct_checker
def svg_dna_transform(pathIN: str, pathOUT: str) -> Tuple[str, str]:
    '''
     ------------------------------------------------------------
    |             TRANSFORM dna_features_viewer SVG              |
    |------------------------------------------------------------|
    |         Transform SVG plot from dna_features_viewer        |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN     : path of input/output SVG file              |
    |    pathOUT    : path of output PNG file (required)         |
     ------------------------------------------------------------
    |TOOLS: convert                                              |
     ------------------------------------------------------------
    '''
    pathIN = path_converter(pathIN)
    pathOUT = path_converter(pathOUT)
    dicoGeminiPath, dicoGeminiModule = get_gemini_path()
    if 'convert' in dicoGeminiModule:
        os.system("module load "+dicoGeminiModule['convert'])

    # ***** DELETE background rectangle and middle line ***** #
    SVG = open(pathIN, 'r')
    lstLines = SVG.read().split("\n")
    SVG.close()
    NEWSVG = open(pathIN, 'w')
    findBg = False
    findHline = False
    for i in range(len(lstLines)):
        if "id = \"patch_1\"" in lstLines[i]:
            findBg = True
        if findBg and "</g>" in lstLines[i]:
            findBg = False
            continue
        if "id = \"line2d_1\"" in lstLines[i]:
            findHline = True
        if findHline and "</g>" in lstLines[i]:
            findHline = False
            continue
        if findBg is False and findHline is False:
            if "<g" in lstLines[i] and "<g" in lstLines[i-1]:
                continue
            NEWSVG.write(lstLines[i]+"\n")
    NEWSVG.close()
    # Parse SVG
    paths, attributes = svg2paths(pathIN)
    # Search max y pos
    maxAllY = 0
    for attr in attributes:
        if "style" in attr:
            searchallY = re.findall(r"([lL]) (\S+) (\S+)\s+Q", attr['d'])
            for searchY in searchallY:
                maxAllY = max(maxAllY, float(searchY[2]))
    # Check required y position transformation
    dicoReplace = {}
    for attr in attributes:
        if "style" in attr:
            maxY = 0
            searchallY = re.findall(r"([lL]) (\S+) (\S+)\s+Q", attr['d'])
            for searchY in searchallY:
                maxY = max(maxY, float(searchY[2]))
            searchDLine = re.search(r"([mM]) (\S+) (\S+)\s+Q", attr['d'])
            if maxY < maxAllY:
                dicoReplace[searchDLine.group(1)+" "+searchDLine.group(2)+" "+searchDLine.group(3)] = maxAllY-float(maxY)
            else:
                dicoReplace[searchDLine.group(1)+" "+searchDLine.group(2)+" "+searchDLine.group(3)] = None
    # Apply y position transformation
    SVG = open(pathIN, 'r')
    lstLines = SVG.read().split("\n")
    SVG.close()
    lstToWrite = []
    for i in range(len(lstLines)):
        for key in dicoReplace:
            if key in lstLines[i]:
                lstToWrite[-1] = lstToWrite[-1].replace(">", "")
                if dicoReplace[key] is None:
                    lstToWrite.append("     style = \"stroke: #000000;stroke-opacity: 1\">")
                else:
                    lstToWrite.append("     style = \"stroke: #000000;stroke-opacity: 1\"")
                    lstToWrite.append("     transform = \"translate(0, "+str(dicoReplace[key])+")\">")
                break
        lstToWrite.append(lstLines[i])
    # Convert to PNG
    svg2png(bytestring="\n".join(lstToWrite), write_to=pathOUT, dpi=300)
    os.system(dicoGeminiPath['TOOLS']['convert']+" "+pathOUT+" -gravity South -background white -splice 0x1 -background black -splice 0x1 -trim +repage -chop 0x1" +
              " -gravity North -background white -splice 0x1 -background black -splice 0x1 -trim +repage -chop 0x1" +
              " -gravity East -background white -splice 5x0 -background black -splice 5x0 -trim +repage -chop 5x " +
              pathOUT.replace(".png", ".bak"))
    shutil.move(pathOUT.replace(".png", ".bak"), pathOUT)


@fct_checker
def xlsx_to_heatmap(pathIN: str, pathOUT: str, colorStart: str = "FFFFFF", colorEnd: str = "FF0000", headerRow: int = -1, headerCol: int = -1) -> Tuple[str, str, str, str, int, int]:
    '''
     ------------------------------------------------------------
    |                      XLSX TO HEATMAP                       |
    |------------------------------------------------------------|
    |            Convert xlsx table to heatmap table             |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN    : path of input files or folder (required)    |
    |    pathOUT   : path of output files (required)             |
    |    colorStart: hex start color (default=FFFFFF)            |
    |    colorEnd  : hex end color (default=FF0000)              |
    |    headerRow : row header line (default=-1 no header)      |
    |    headerCol : column header line (default=-1 no header)   |
     ------------------------------------------------------------
    '''
    pathIN = path_converter(pathIN)
    pathOUT = path_converter(pathOUT)
    printcolor("♊ XLSX to heatmap"+"\n")
    # Create color gradient list
    HEX_list, RBG_list = linear_gradient("#"+colorStart, "#"+colorEnd, 101)
    # Open input/output XLSX
    spinner = yaspin(Spinners.aesthetic, text="♊ Loading input XLSX", side="right")
    spinner.start()
    title("Loading", None)
    wbIN = load_workbook(pathIN)
    wsIN = wbIN.active
    spinner.stop()
    printcolor("♊ Load XLSX"+"\n")
    wbOUT = Workbook(write_only=True)
    wsOUT = wbOUT.create_sheet("heatmap")
    # Browse input XLSX
    printcolor("♊ Heatmap creation"+"\n")
    cptRow = 0
    pbar = tqdm(total=wsIN.max_row, dynamic_ncols=True, ncols=50, leave=False, desc="", file=sys.stdout, bar_format="  {percentage: 3.0f}%|{bar}| {n_fmt}/{total_fmt}")
    for rowIN in wsIN.rows:
        rowOUT = []
        cptCol = 0
        for cellIN in rowIN:
            # For header copy value
            if cptRow == headerRow or cptCol == headerCol:
                cellOUT = WriteOnlyCell(wsOUT, value=cellIN.value)
            # For data apply color without value
            else:
                cellOUT = WriteOnlyCell(wsOUT, value="")
                value = int(round(float(str(cellIN.value).replace(", ", ".")), 0))
                cellColor = styles.colors.Color(rgb=HEX_list[value].replace("#", ""))
                if cellColor != "#FFFFFF":
                    cellFill = styles.fills.PatternFill(patternType='solid', fgColor=cellColor)
                    cellOUT.fill = cellFill
            cellOUT.alignment = Alignment(horizontal='center')
            rowOUT.append(cellOUT)
            cptCol += 1
        wsOUT.append(rowOUT)
        cptRow += 1
        pbar.update(1)
        title("Heatmap", pbar)
    pbar.close()
    # Close the workbook after reading
    wbIN.close()
    wbOUT.save(pathOUT)


@fct_checker
def circos_plot(pathIN1: str, pathOUT: str, pathIN2: str = "None", pident: int = 30, cov: int = 80) -> Tuple[str, str, str, int, int]:
    '''
     ------------------------------------------------------------
    |                        CIRCOS PLOT                         |
    |------------------------------------------------------------|
    |           Circular circos plot from genbank file           |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN1 : path of reference genbank file (required)     |
    |    pathOUT : path of output file (required)                |
    |    pathIN2 : path of query FAA file/folder (default=None)  |
    |    pident  : min %id for orthologuous (default=30)         |
    |    cov     : min %cov for orthologuous (default=80)        |
     ------------------------------------------------------------
    |TOOLS: circos                                               |
     ------------------------------------------------------------
    '''
    pathIN1 = path_converter(pathIN1)
    pathOUT = path_converter(pathOUT)
    dicoGeminiPath, dicoGeminiModule = get_gemini_path()
    if 'circos' in dicoGeminiModule:
        os.system("module load "+dicoGeminiModule['circos'])
    if pathIN2 != "None":
        lstQueryFAAFiles, maxpathSize = get_input_files(pathIN2, "circos_plot", [".faa"])

    # ***** READ reference GBK ***** #
    printcolor("♊ Parse GBK"+"\n")
    dicoGBK = list(make_gbk_dict(pathIN1).values())[0]

    # ***** BLASTP query FAA ***** #
    if pathIN2 != "None":
        # Init dicoRBH
        dicoRBH = {}
        for contig in dicoGBK:
            for lt in dicoGBK[contig]['dicoLT']:
                dicoRBH[lt] = {}
                for FAAFile in lstQueryFAAFiles:
                    dicoRBH[lt][os.path.basename(FAAFile).replace(".faa", "")] = 0
        # Make FAA and create symlinks
        gbk_to_faa(pathIN=pathIN1, pathOUT=geminiset.pathTMP+"/ref.faa")
        for FAAFile in lstQueryFAAFiles:
            orgQuery = os.path.basename(FAAFile).replace(".faa", "")
            os.makedirs(geminiset.pathTMP+"/mmseqs")
            os.symlink(FAAFile, geminiset.pathTMP+"/mmseqs/"+os.path.basename(FAAFile))
            os.symlink(geminiset.pathTMP+"/ref.faa", geminiset.pathTMP+"/mmseqs/ref.faa")
            # Launch mmseqs_rbh
            mmseqs_rbh(pathIN=geminiset.pathTMP+"/mmseqs", pathOUT=geminiset.pathTMP+"/mmseqs", ref="ref", idThrClust=pident, covThrClust=cov, boolNucl=False, ext=".faa")
            shutil.copyfile(geminiset.pathTMP+"/mmseqs/ref.rbh", geminiset.pathTMP+"/"+orgQuery+".rbh")
            os.system("rm -rf "+geminiset.pathTMP+"/mmseqs*")
            # Parse mmseqs_rbh
            TSV = open(geminiset.pathTMP+"/"+orgQuery+".rbh", 'r')
            lstLines = TSV.read().split("\n")[:-1]
            TSV.close()
            for line in lstLines:
                splitLine = line.split("\t")
                ltRef = splitLine[0].replace("#", "|").split("|")[0]
                dicoRBH[ltRef][orgQuery] = float(splitLine[2])

    # ***** KARYOTYPE ***** #
    KARYOTYPE = open(geminiset.pathTMP+"/karyotype.txt", 'w')
    lstColorBand = ["black", "white", "grey"]
    start = 0
    bandIndex = 0
    totalSize = 0
    LstTowrite = []
    for contig in dicoGBK:
        ctgLen = len(dicoGBK[contig]['seq'])
        totalSize += ctgLen
        LstTowrite.append("band chr1 1.1 1.1 "+str(start)+" "+str(start+ctgLen)+" "+lstColorBand[bandIndex])
        bandIndex += 1
        if bandIndex == 3:
            bandIndex = 0
        start = start+ctgLen
    # Add total size at first karyotype line
    LstTowrite.insert(0, "chr - chr1 1 0 "+str(totalSize)+" black")
    KARYOTYPE.write("\n".join(LstTowrite))
    KARYOTYPE.close()
    # Define ticks correspunding to total size
    if totalSize >= 1000000:
        ticksMultiplier = 0.000001
        majorTicksSpacing = "100000u"
        minorTicksSpacing = "10000u"
        majorTicksSuffix = "Mb"
    elif totalSize >= 1000:
        ticksMultiplier = 0.001
        majorTicksSpacing = "1000u"
        minorTicksSpacing = "100u"
        majorTicksSuffix = "Kb"
    # ***** IDEOGRAM ***** #
    IDEOGRAM = open(geminiset.pathTMP+"/ideogram.conf", 'w')
    IDEOGRAM.write("<ideogram>\n")
    IDEOGRAM.write("<spacing>\n")
    IDEOGRAM.write("default = 0.001r\n")
    IDEOGRAM.write("break   = 50r\n")
    IDEOGRAM.write("</spacing>\n")
    IDEOGRAM.write("thickness             = 20p\n")  # thickness (px) of chromosome ideogram
    IDEOGRAM.write("stroke_thickness      = 2\n")
    IDEOGRAM.write("stroke_color          = black\n")  # ideogram border color
    IDEOGRAM.write("fill                  = yes\n")
    IDEOGRAM.write("fill_color            = black\n")  # the default chromosome color is set here and any value, defined in the karyotype file overrides it
    IDEOGRAM.write("radius                = 0.85r\n")  # fractional radius position of chromosome ideogram within image
    IDEOGRAM.write("show_label            = no\n")
    IDEOGRAM.write("label_font            = default\n")
    IDEOGRAM.write("label_radius          = dims(ideogram,radius) + 0.05r\n")
    IDEOGRAM.write("label_size            = 36\n")
    IDEOGRAM.write("label_parallel        = yes\n")
    IDEOGRAM.write("label_case            = upper\n")
    IDEOGRAM.write("band_stroke_thickness = 2\n")  # cytogenetic bands
    IDEOGRAM.write("show_bands            = yes\n")  # show_bands determines whether the outline of cytogenetic bands will be seen
    IDEOGRAM.write("fill_bands            = yes\n")  # in order to fill the bands with the color defined in the karyotype file you must set fill_bands
    IDEOGRAM.write("</ideogram>\n")
    IDEOGRAM.close()

    # ***** TICKS ***** #
    TICKS = open(geminiset.pathTMP+"/ticks.conf", 'w')
    TICKS.write("show_ticks          = yes\n")
    TICKS.write("show_tick_labels    = yes\n")
    TICKS.write("show_grid          = no\n")
    TICKS.write("grid_start         = dims(ideogram,radius_inner)-0.5r\n")
    TICKS.write("grid_end           = dims(ideogram,radius_inner)\n")
    TICKS.write("<ticks>\n")
    TICKS.write("skip_first_label     = yes\n")
    TICKS.write("skip_last_label      = no\n")
    TICKS.write("radius               = dims(ideogram,radius_outer)\n")
    TICKS.write("tick_separation      = 2p\n")
    TICKS.write("min_label_distance_to_edge = 0p\n")
    TICKS.write("label_separation = 5p\n")
    TICKS.write("label_offset     = 25p\n")
    TICKS.write("label_size = 8p\n")
    TICKS.write("multiplier = "+str(ticksMultiplier)+"\n")
    TICKS.write("color = black\n")
    # Major ticks
    TICKS.write("<tick>\n")
    TICKS.write("spacing        = "+majorTicksSpacing+"\n")
    TICKS.write("color          = black\n")
    TICKS.write("show_label     = yes\n")
    TICKS.write("suffix = \" "+majorTicksSuffix+"\"\n")
    TICKS.write("label_size     = 36p\n")
    TICKS.write("format         = %s\n")
    TICKS.write("grid           = yes\n")
    TICKS.write("grid_color     = dgrey\n")
    TICKS.write("grid_thickness = 1p\n")
    TICKS.write("thickness = 5p\n")
    TICKS.write("size      = 30p\n")
    TICKS.write("</tick>\n")
    # Minor ticks
    TICKS.write("<tick>\n")
    TICKS.write("spacing        = "+minorTicksSpacing+"\n")
    TICKS.write("color          = grey\n")
    TICKS.write("show_label     = no\n")
    TICKS.write("format         = %s\n")
    TICKS.write("grid           = yes\n")
    TICKS.write("grid_color     = dgrey\n")
    TICKS.write("grid_thickness = 1p\n")
    TICKS.write("thickness = 5p\n")
    TICKS.write("size      = 20p\n")
    TICKS.write("</tick>\n")
    TICKS.write("</ticks>\n")
    TICKS.close()

    # ***** COLORS ***** #
    COLORS = open(geminiset.pathTMP+"/colors.conf", 'w')
    COLORS.write("generev = 77,77,77\n")
    COLORS.write("genefor = 179,179,179\n")
    COLORS.write("generna = 0,0,0\n")
    # Gradient
    HEXlist, RBGlist = linear_gradient("#db073d", "#ffe9e7", 100)
    cpt = 1
    strLstGradient = ""
    for rgb in RBGlist[::-1]:
        COLORS.write("gradient"+str(cpt)+" = "+str(rgb[0])+","+str(rgb[1])+","+str(rgb[2])+"\n")
        strLstGradient += "gradient"+str(cpt)+","
        cpt += 1
    COLORS.close()
    strLstGradient = strLstGradient[:-1]

    # ***** HIGHLIGHTS ***** #
    HIGHLIGHTSFOR = open(geminiset.pathTMP+"/highlightsFOR.txt", 'w')
    totalSize = 0
    for contig in dicoGBK:
        ctgLen = len(dicoGBK[contig]['seq'])
        for lt in dicoGBK[contig]['dicoLT']:
            if dicoGBK[contig]['dicoLT'][lt]['type'] == "CDS" and dicoGBK[contig]['dicoLT'][lt]['strand'] == 1:
                HIGHLIGHTSFOR.write("chr1 "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['start'])+" "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['end'])+" fill_color=genefor\n")
        totalSize += ctgLen
    HIGHLIGHTSFOR.close()
    HIGHLIGHTSREV = open(geminiset.pathTMP+"/highlightsREV.txt", 'w')
    totalSize = 0
    for contig in dicoGBK:
        ctgLen = len(dicoGBK[contig]['seq'])
        for lt in dicoGBK[contig]['dicoLT']:
            if dicoGBK[contig]['dicoLT'][lt]['type'] == "CDS" and dicoGBK[contig]['dicoLT'][lt]['strand'] == -1:
                HIGHLIGHTSREV.write("chr1 "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['start'])+" "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['end'])+" fill_color=generev\n")
        totalSize += ctgLen
    HIGHLIGHTSREV.close()
    HIGHLIGHTSRNA = open(geminiset.pathTMP+"/highlightsRNA.txt", 'w')
    totalSize = 0
    for contig in dicoGBK:
        ctgLen = len(dicoGBK[contig]['seq'])
        for lt in dicoGBK[contig]['dicoLT']:
            if "RNA" in dicoGBK[contig]['dicoLT'][lt]['type']:
                HIGHLIGHTSRNA.write("chr1 "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['start'])+" "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['end'])+" fill_color=generna\n")
        totalSize += ctgLen
    HIGHLIGHTSRNA.close()

    # ***** ORTHOLOGOUS HEATMAP ***** #
    totalSize = 0
    if pathIN2 != "None":
        for contig in dicoGBK:
            ctgLen = len(dicoGBK[contig]['seq'])
            for lt in dicoGBK[contig]['dicoLT']:
                for orgQuery in dicoRBH[lt]:
                    ORTHOORG = open(geminiset.pathTMP+"/"+orgQuery+"_heatmap.txt", 'a')
                    ORTHOORG.write("chr1 "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['start'])+" "+str(totalSize+dicoGBK[contig]['dicoLT'][lt]['end'])+" "+str(dicoRBH[lt][orgQuery])+"\n")
                    ORTHOORG.close()
            totalSize += ctgLen

    # ***** CIRCOSCONF ***** #
    CIRCOSCONF = open(geminiset.pathTMP+"/circos.conf", 'w')
    CIRCOSCONF.write("karyotype = "+geminiset.pathTMP+"/karyotype.txt\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/ideogram.conf>>\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/ticks.conf>>\n")
    CIRCOSCONF.write("<image>\n")
    CIRCOSCONF.write("dir = "+os.path.realpath(os.path.dirname(pathOUT))+"\n")
    CIRCOSCONF.write("file = "+os.path.basename(pathOUT)+"\n")
    if ".png" in os.path.basename(pathOUT):
        CIRCOSCONF.write("png = yes\n")
        CIRCOSCONF.write("svg = no\n")
    else:
        CIRCOSCONF.write("png = no\n")
        CIRCOSCONF.write("svg = yes\n")
    CIRCOSCONF.write("radius = 1500p\n")
    CIRCOSCONF.write("angle_offset = -82\n")
    CIRCOSCONF.write("auto_alpha_colors = yes\n")
    CIRCOSCONF.write("auto_alpha_steps  = 5\n")
    CIRCOSCONF.write("</image>\n")
    CIRCOSCONF.write("<colors>\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/colors.conf>>\n")
    CIRCOSCONF.write("</colors>\n")
    CIRCOSCONF.write("<highlights>\n")
    CIRCOSCONF.write("<highlight>\n")
    CIRCOSCONF.write("file = "+geminiset.pathTMP+"/highlightsFOR.txt\n")
    CIRCOSCONF.write("r0   = 0.94r\n")
    CIRCOSCONF.write("r1   = 0.99r\n")
    CIRCOSCONF.write("</highlight>\n")
    CIRCOSCONF.write("<highlight>\n")
    CIRCOSCONF.write("file = "+geminiset.pathTMP+"/highlightsREV.txt\n")
    CIRCOSCONF.write("r0   = 0.89r\n")
    CIRCOSCONF.write("r1   = 0.94r\n")
    CIRCOSCONF.write("</highlight>\n")
    CIRCOSCONF.write("<highlight>\n")
    CIRCOSCONF.write("file = "+geminiset.pathTMP+"/highlightsRNA.txt\n")
    CIRCOSCONF.write("r0   = 0.84r\n")
    CIRCOSCONF.write("r1   = 0.89r\n")
    CIRCOSCONF.write("</highlight>\n")
    CIRCOSCONF.write("</highlights>\n")
    if pathIN2 != "None":
        CIRCOSCONF.write("<plots>\n")
        rStart = 0.83
        for file in os.listdir(geminiset.pathTMP):
            if "_heatmap.txt" in file:
                CIRCOSCONF.write("<plot>\n")
                CIRCOSCONF.write("type    = heatmap\n")
                CIRCOSCONF.write("file    = "+geminiset.pathTMP+"/"+file+"\n")
                CIRCOSCONF.write("color   = "+strLstGradient+"\n")
                CIRCOSCONF.write("r0      = "+str(round(rStart-0.02, 2))+"r\n")
                CIRCOSCONF.write("r1      = "+str(rStart)+"r\n")
                CIRCOSCONF.write("min     = "+str(pident)+"\n")
                CIRCOSCONF.write("max     = 100\n")
                CIRCOSCONF.write("</plot>\n")
                rStart = round(rStart-0.03, 2)
        CIRCOSCONF.write("</plots>\n")
    CIRCOSCONF.write("<<include /etc/circos/colors_fonts_patterns.conf>>\n")
    CIRCOSCONF.write("<<include /etc/circos/housekeeping.conf>>\n")
    CIRCOSCONF.close()

    # ***** MAKE PLOT *****#
    print("circos -conf "+geminiset.pathTMP+"/circos.conf")
    os.system(dicoGeminiPath['TOOLS']['circos']+" -conf "+geminiset.pathTMP+"/circos.conf")
    print("\n".join(lstQueryFAAFiles))


@fct_checker
def circos_align(pathIN: str, pathOUT: str) -> Tuple[str, str]:
    '''
     ------------------------------------------------------------
    |                   CIRCOS ALIGNMENT PLOT                    |
    |------------------------------------------------------------|
    |             Circular circos plot for alignment             |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN : path of fasta folder (required)                |
    |    pathOUT : path of output file (required)                |
     ------------------------------------------------------------
    |TOOLS: circos                                               |
     ------------------------------------------------------------
    '''
    dicoGeminiPath, dicoGeminiModule = get_gemini_path()
    if 'circos' in dicoGeminiModule:
        os.system("module load "+dicoGeminiModule['circos'])
    pathIN = path_converter(pathIN)
    pathOUT = path_converter(pathOUT)
    lstFiles, maxpathSize = get_input_files(pathIN, "circos_align", [".fna", ".fasta"])

    # ***** CAT files ***** #
    cat_lstfiles(lstFiles, geminiset.pathTMP+"/catseq.fasta")

    # ***** LAUNCH MAFFT ***** #
    spinner = yaspin(Spinners.aesthetic, text="♊ MAFFT", side="right")
    spinner.start()
    title("MAFFT", None)
    cmdMAFFT = dicoGeminiPath['TOOLS']['mafft']+" --quiet --thread "+str(geminiset.cpu)+" "+geminiset.pathTMP+"/catseq.fasta > "+geminiset.pathTMP+"/mafft.fasta"
    os.system(cmdMAFFT)
    spinner.stop()
    printcolor("♊ MAFFT"+"\n")

    # ***** PARSE MAFFT ***** #
    dicoFASTA = make_fasta_dict(geminiset.pathTMP+"/mafft.fasta")
    alignLen = len(list(dicoFASTA.values())[0])
    dicoMAFFT = {}
    for align in dicoFASTA:
        for i in range(0, alignLen, 1):
            if dicoFASTA[align][i] != "N":
                try:
                    dicoMAFFT[i].add(dicoFASTA[align][i])
                except KeyError:
                    dicoMAFFT[i] = set([dicoFASTA[align][i]])

    # ***** KARYOTYPE ***** #
    KARYOTYPE = open(geminiset.pathTMP+"/karyotype.txt", 'w')
    KARYOTYPE.write("chr - chr1 1 0 "+str(alignLen)+" black\n")
    KARYOTYPE.write("chr - chr2 1 "+str(alignLen+1)+" "+str(alignLen+2)+" white\n")
    KARYOTYPE.close()
    # Define ticks correspunding to total size
    if alignLen >= 4000000:
        ticksMultiplier = 0.000001
        majorTicksSpacing = "1000000u"
        minorTicksSpacing = "100000u"
        majorTicksSuffix = "Mb"
    elif alignLen >= 100000:
        ticksMultiplier = 0.00001
        majorTicksSpacing = "100000u"
        minorTicksSpacing = "10000u"
        majorTicksSuffix = "Kb"

    # ****** COLORS *****#
    COLORS = open(geminiset.pathTMP+"/colors.conf", 'w')
    COLORS.write("cons = 175,221,233\n")
    COLORS.write("diff = 255,85,85\n")
    COLORS.write("gap  = 255,255,255\n")
    COLORS.close()

    # ***** IDEOGRAM ***** #
    IDEOGRAM = open(geminiset.pathTMP+"/ideogram.conf", 'w')
    IDEOGRAM.write("<ideogram>\n")
    IDEOGRAM.write("<spacing>\n")
    IDEOGRAM.write("default = 0u\n")
    IDEOGRAM.write("break   = 0u\n")
    IDEOGRAM.write("<pairwise chr2 chr1>\n")
    IDEOGRAM.write("spacing = "+str(int(alignLen/33))+"u\n")
    IDEOGRAM.write("</pairwise>\n")
    IDEOGRAM.write("</spacing>\n")
    IDEOGRAM.write("thickness             = 20p\n")  # thickness (px) of chromosome ideogram
    IDEOGRAM.write("fill                  = yes\n")
    IDEOGRAM.write("fill_color            = black\n")  # the default chromosome color is set here and any value, defined in the karyotype file overrides it
    IDEOGRAM.write("radius                = 0.85r\n")  # fractional radius position of chromosome ideogram within image
    IDEOGRAM.write("show_label            = no\n")
    IDEOGRAM.write("label_font            = default\n")
    IDEOGRAM.write("label_radius          = dims(ideogram,radius) + 0.05r\n")
    IDEOGRAM.write("label_size            = 36\n")
    IDEOGRAM.write("label_parallel        = yes\n")
    IDEOGRAM.write("label_case            = upper\n")
    IDEOGRAM.write("band_stroke_thickness = 2\n")  # cytogenetic bands
    IDEOGRAM.write("show_bands            = yes\n")  # show_bands determines whether the outline of cytogenetic bands will be seen
    IDEOGRAM.write("fill_bands            = yes\n")  # in order to fill the bands with the color defined in the karyotype file you must set fill_bands
    IDEOGRAM.write("</ideogram>\n")
    IDEOGRAM.close()

    # ***** TICKS ***** #
    TICKS = open(geminiset.pathTMP+"/ticks.conf", 'w')
    TICKS.write("show_ticks          = yes\n")
    TICKS.write("show_tick_labels    = yes\n")
    TICKS.write("show_grid          = no\n")
    TICKS.write("grid_start         = dims(ideogram,radius_inner)-0.5r\n")
    TICKS.write("grid_end           = dims(ideogram,radius_inner)\n")
    TICKS.write("<ticks>\n")
    TICKS.write("skip_first_label     = yes\n")
    TICKS.write("skip_last_label      = no\n")
    TICKS.write("radius               = dims(ideogram,radius_outer)\n")
    TICKS.write("tick_separation      = 2p\n")
    TICKS.write("min_label_distance_to_edge = 0p\n")
    TICKS.write("label_separation = 5p\n")
    TICKS.write("label_offset     = 25p\n")
    TICKS.write("label_size = 8p\n")
    TICKS.write("multiplier = "+str(ticksMultiplier)+"\n")
    TICKS.write("color = black\n")
    # Major ticks
    TICKS.write("<tick>\n")
    TICKS.write("spacing        = "+majorTicksSpacing+"\n")
    TICKS.write("color          = black\n")
    TICKS.write("show_label     = yes\n")
    TICKS.write("suffix = \" "+majorTicksSuffix+"\"\n")
    TICKS.write("label_size     = 36p\n")
    TICKS.write("format         = %s\n")
    TICKS.write("grid           = yes\n")
    TICKS.write("grid_color     = dgrey\n")
    TICKS.write("grid_thickness = 1p\n")
    TICKS.write("thickness = 5p\n")
    TICKS.write("size      = 30p\n")
    TICKS.write("chromosomes = -chr2\n")
    TICKS.write("</tick>\n")
    # Minor ticks
    TICKS.write("<tick>\n")
    TICKS.write("spacing        = "+minorTicksSpacing+"\n")
    TICKS.write("color          = grey\n")
    TICKS.write("show_label     = no\n")
    TICKS.write("format         = %s\n")
    TICKS.write("grid           = yes\n")
    TICKS.write("grid_color     = dgrey\n")
    TICKS.write("grid_thickness = 1p\n")
    TICKS.write("thickness = 5p\n")
    TICKS.write("size      = 20p\n")
    TICKS.write("chromosomes = -chr2\n")
    TICKS.write("</tick>\n")
    TICKS.write("</ticks>\n")
    TICKS.close()

    # ***** HIGHLIGHTS ***** #
    cpt = 1
    lstHighlights = []
    dicoColor = {"X": "diff", "-": "gap", "*": "cons"}
    for align in dicoFASTA:
        dicoGroupByColor = {"*": [], "X": [], "-": []}
        lstHighlights.append(geminiset.pathTMP+"/highlights"+str(cpt)+".txt")
        for i in range(0, alignLen, 1):
            nucl = dicoFASTA[align][i].upper().replace("A", "X").replace("T", "X").replace("G", "X").replace("C", "X").replace("N", "X")
            if i not in dicoMAFFT or len(dicoMAFFT[i]) == 1:
                dicoGroupByColor["*"].append(i)
            else:
                dicoGroupByColor[nucl].append(i)
        # Range per organism and by color
        HIGHLIGHTSALIGN = open(geminiset.pathTMP+"/highlights"+str(cpt)+".txt", 'w')
        for nuclType in dicoGroupByColor:
            lstGrpByType = list(to_ranges(dicoGroupByColor[nuclType]))
            for grpByType in lstGrpByType:
                HIGHLIGHTSALIGN.write("chr1 "+str(grpByType[0])+" "+str(grpByType[1])+" fill_color="+dicoColor[nuclType]+"\n")
        HIGHLIGHTSALIGN.close()
        cpt += 1

    # ***** CIRCOSCONF ***** #
    CIRCOSCONF = open(geminiset.pathTMP+"/circos.conf", 'w')
    CIRCOSCONF.write("karyotype = "+geminiset.pathTMP+"/karyotype.txt\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/ideogram.conf>>\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/ticks.conf>>\n")
    CIRCOSCONF.write("<image>\n")
    CIRCOSCONF.write("dir = "+os.path.realpath(os.path.dirname(pathOUT))+"\n")
    CIRCOSCONF.write("file = "+os.path.basename(pathOUT)+"\n")
    CIRCOSCONF.write("png = yes\n")
    CIRCOSCONF.write("svg = no\n")
    CIRCOSCONF.write("radius = 1500p\n")
    CIRCOSCONF.write("angle_offset = -80\n")
    CIRCOSCONF.write("auto_alpha_colors = yes\n")
    CIRCOSCONF.write("auto_alpha_steps  = 5\n")
    CIRCOSCONF.write("</image>\n")
    CIRCOSCONF.write("<colors>\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/colors.conf>>\n")
    CIRCOSCONF.write("</colors>\n")
    CIRCOSCONF.write("<highlights>\n")
    rstart = 0.99
    for highlights in lstHighlights:
        CIRCOSCONF.write("<highlight>\n")
        CIRCOSCONF.write("file = "+highlights+"\n")
        CIRCOSCONF.write("r0   = "+str(round(rstart-0.04, 2))+"r\n")
        CIRCOSCONF.write("r1   = "+str(rstart)+"r\n")
        CIRCOSCONF.write("</highlight>\n")
        rstart = round(rstart-0.05, 2)
    CIRCOSCONF.write("</highlights>\n")
    CIRCOSCONF.write("<<include /etc/circos/colors_fonts_patterns.conf>>\n")
    CIRCOSCONF.write("<<include /etc/circos/housekeeping.conf>>\n")
    CIRCOSCONF.close()

    # ***** MAKE PLOT *****#
    os.system("echo \"circos -conf "+geminiset.pathTMP+"/circos.conf\" > "+geminiset.pathTMP+"/circos.log")
    os.system(dicoGeminiPath['TOOLS']['circos']+" -conf "+geminiset.pathTMP+"/circos.conf >> "+geminiset.pathTMP+"/circos.log 2>&1")
    # Display sequence name order
    print("Order from outside to inside: ")
    for align in dicoFASTA:
        print(align)
    exit()


@fct_checker
def circos_rbh_plot(pathIN: str, pathJSON: str, pathOUT: str, maxCore: float = 0.9) -> Tuple[str, str, str, int]:
    '''
     ------------------------------------------------------------
    |                       CIRCOS RBH PLOT                      |
    |------------------------------------------------------------|
    |    Circular circos plot from genbank file with rbh link    |
    |------------------------------------------------------------|
    |PARAMETERS                                                  |
    |    pathIN   : path of genbank files (required)             |
    |    pathJSON : path of rbh clusters JSON file (required)    |
    |    pathOUT  : path of output file (required)               |
    |    freq     : max core frequence to show (default=0.9)     |
     ------------------------------------------------------------
    |TOOLS: circos                                               |
     ------------------------------------------------------------
    '''
    pathIN = path_converter(pathIN)
    pathJSON = path_converter(pathJSON)
    pathOUT = path_converter(pathOUT)
    lstGBKFiles, maxpathSize = get_input_files(pathIN, "circos_rbh_plot", [".gbk", ".gbk.gz"])
    dicoGeminiPath, dicoGeminiModule = get_gemini_path()
    if 'circos' in dicoGeminiModule:
        os.system("module load "+dicoGeminiModule['circos'])

    # ***** READ GBKs & KARYOTYPE ***** #
    printcolor("♊ Parse GBK"+"\n")
    dicoGBK = {}
    dicoGBKLT = {}
    dicoOrgToId = {}
    KARYOTYPE = open(geminiset.pathTMP+"/karyotype.txt", 'w')
    lstColorBand = ["black", "white", "grey"]
    cptOrg = 1
    for gbkFile in lstGBKFiles:
        orgName = os.path.basename(gbkFile).replace(".gbk.gz", "").replace(".gbk", "").replace("Vibrio_crassostreae_", "").replace("_chr1", "")
        dicoGBK[orgName] = list(make_gbk_dict(gbkFile).values())[0]
        cptCtg = 1
        start = 0
        totalSize = 0
        lstBand = []
        bandIndex = 0
        for contig in dicoGBK[orgName]:
            ctgLen = len(dicoGBK[orgName][contig]['seq'])
            totalSize += ctgLen
            lstBand.append("band chr"+str(cptOrg)+" "+str(cptOrg)+"."+str(cptCtg)+" "+str(cptOrg)+"."+str(cptCtg)+" "+str(start)+" "+str(start+ctgLen)+" "+lstColorBand[bandIndex])
            bandIndex += 1
            if bandIndex == 3:
                bandIndex = 0
            for lt in dicoGBK[orgName][contig]['dicoLT']:  # {'type', 'product', 'protSeq', 'start', 'end', 'strand', 'protein_id', 'geneSeq'}
                if dicoGBK[orgName][contig]['dicoLT'][lt]['type'] == "CDS":
                    shortlt = lt.split(" ")[0]
                    dicoGBKLT[shortlt] = {'relstart': start+dicoGBK[orgName][contig]['dicoLT'][lt]['start'], 'relend': start+dicoGBK[orgName][contig]['dicoLT'][lt]['end']}
            start = start+ctgLen
            cptCtg += 1
        KARYOTYPE.write("chr - chr"+str(cptOrg)+" "+orgName+" 0 "+str(totalSize)+" black\n")
        KARYOTYPE.write("\n".join(lstBand)+"\n")
        dicoOrgToId[orgName] = "chr"+str(cptOrg)
        cptOrg += 1
    KARYOTYPE.close()

    # ***** IDEOGRAM ***** #
    printcolor("♊ Make Ideogram"+"\n")
    IDEOGRAM = open(geminiset.pathTMP+"/ideogram.conf", 'w')
    IDEOGRAM.write("<ideogram>\n")
    IDEOGRAM.write("<spacing>\n")
    IDEOGRAM.write("default = 100000u\n")
    IDEOGRAM.write("break   = 0u\n")
    IDEOGRAM.write("</spacing>\n")
    IDEOGRAM.write("thickness             = 20p\n")  # thickness (px) of chromosome ideogram
    IDEOGRAM.write("stroke_thickness      = 2\n")
    IDEOGRAM.write("stroke_color          = black\n")  # ideogram border color
    IDEOGRAM.write("fill                  = yes\n")
    IDEOGRAM.write("fill_color            = black\n")  # the default chromosome color is set here and any value, defined in the karyotype file overrides it
    IDEOGRAM.write("radius                = 0.85r\n")  # fractional radius position of chromosome ideogram within image
    IDEOGRAM.write("show_label            = yes\n")
    IDEOGRAM.write("label_font            = default\n")
    IDEOGRAM.write("label_radius          = dims(ideogram,radius) + 0.05r\n")
    IDEOGRAM.write("label_size            = 36\n")
    IDEOGRAM.write("label_parallel        = yes\n")
    IDEOGRAM.write("label_case            = upper\n")
    IDEOGRAM.write("band_stroke_thickness = 2\n")  # cytogenetic bands
    IDEOGRAM.write("show_bands            = yes\n")  # show_bands determines whether the outline of cytogenetic bands will be seen
    IDEOGRAM.write("fill_bands            = yes\n")  # in order to fill the bands with the color defined in the karyotype file you must set fill_bands
    IDEOGRAM.write("</ideogram>\n")
    IDEOGRAM.close()

    # ****** COLORS ***** #
    COLORS = open(geminiset.pathTMP+"/colors.conf", 'w')
    # Gradient
    HEXlist, RBGlist = linear_gradient("#ffffff", "#db073d", len(dicoGBK))
    cpt = 1
    strLstGradient = ""
    for rgb in RBGlist[::-1]:
        COLORS.write("gradient"+str(cpt)+" = "+str(rgb[0])+","+str(rgb[1])+","+str(rgb[2])+"\n")
        strLstGradient += "gradient"+str(cpt)+","
        cpt += 1
    COLORS.close()
    strLstGradient = strLstGradient[:-1]

    # ***** RBH LINKS & HEATMAP ***** #
    printcolor("♊ Make RBH links / heatmap"+"\n")
    dicoRBH = load_json(pathJSON)
    LINKS = open(geminiset.pathTMP+"/links1.txt", 'w')
    cptLink = 1
    cptLinkPerFile = 0
    cptLinkFile = 1
    for clusterNum in dicoRBH:
        setOrg = set()
        for prot in dicoRBH[clusterNum]:
            print("#"+prot+"#")
            setOrg.add(prot.split("[")[1].replace("]", ""))
        if len(setOrg) > 1:
            setCompDone = set()
            for prot1 in dicoRBH[clusterNum]:
                lt1 = prot1.split(" ")[0]
                orgName1 = prot1.split("[")[1].replace("]", "")
                # Heatmap
                HEATMAP = open(geminiset.pathTMP+"/heatmap_"+orgName1+".txt", 'a')
                HEATMAP.write(dicoOrgToId[orgName1]+" "+str(dicoGBKLT[lt1]['relstart'])+" "+str(dicoGBKLT[lt1]['relend'])+" "+str(len(setOrg))+"\n")
                HEATMAP.close()
                if len(setOrg)/len(dicoGBK) <= maxCore:
                    for prot2 in dicoRBH[clusterNum]:
                        lt2 = prot2.split(" ")[0]
                        orgName2 = prot2.split("[")[1].replace("]", "")
                        if orgName1 != orgName2 and orgName1+"#"+orgName2 not in setCompDone and orgName2+"#"+orgName1 not in setCompDone:
                            LINKS.write("link"+str(cptLink)+" "+dicoOrgToId[orgName1]+" "+str(dicoGBKLT[lt1]['relstart'])+" "+str(dicoGBKLT[lt1]['relend'])+" color=gradient"+str(len(setOrg))+"\n")
                            LINKS.write("link"+str(cptLink)+" "+dicoOrgToId[orgName2]+" "+str(dicoGBKLT[lt2]['relstart'])+" "+str(dicoGBKLT[lt2]['relend'])+" color=gradient"+str(len(setOrg))+"\n")
                            cptLink += 1
                            cptLinkPerFile += 2
                            if cptLinkPerFile > 10000:
                                LINKS.close()
                                cptLinkPerFile = 0
                                cptLinkFile += 1
                                LINKS = open(geminiset.pathTMP+"/links"+str(cptLinkFile)+".txt", 'w')
                            setCompDone.add(orgName1+"#"+orgName2)
                            setCompDone.add(orgName2+"#"+orgName1)
    LINKS.close()

    # ***** CIRCOSCONF ***** #
    CIRCOSCONF = open(geminiset.pathTMP+"/circos.conf", 'w')
    CIRCOSCONF.write("karyotype = "+geminiset.pathTMP+"/karyotype.txt\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/ideogram.conf>>\n")
    CIRCOSCONF.write("<image>\n")
    CIRCOSCONF.write("dir = "+os.path.realpath(os.path.dirname(pathOUT))+"\n")
    CIRCOSCONF.write("file = "+os.path.basename(pathOUT)+"\n")
    if ".png" in os.path.basename(pathOUT):
        CIRCOSCONF.write("png = yes\n")
        CIRCOSCONF.write("svg = no\n")
    else:
        CIRCOSCONF.write("png = no\n")
        CIRCOSCONF.write("svg = yes\n")
    CIRCOSCONF.write("radius = 1500p\n")
    CIRCOSCONF.write("angle_offset = -90\n")
    CIRCOSCONF.write("auto_alpha_colors = yes\n")
    CIRCOSCONF.write("auto_alpha_steps  = 5\n")
    CIRCOSCONF.write("</image>\n")
    CIRCOSCONF.write("<colors>\n")
    CIRCOSCONF.write("<<include "+geminiset.pathTMP+"/colors.conf>>\n")
    CIRCOSCONF.write("</colors>\n")
    # Links
    CIRCOSCONF.write("<links>\n")
    for file in os.listdir(geminiset.pathTMP):
        if "links" in file:
            CIRCOSCONF.write("<link>\n")
            CIRCOSCONF.write("file          = "+geminiset.pathTMP+"/"+file+"\n")
            CIRCOSCONF.write("radius        = 0.90r\n")
            # Curves look best when this value is small (e.g. 0.1r or 0r)
            CIRCOSCONF.write("bezier_radius = 0.1r\n")
            CIRCOSCONF.write("thickness     = 1\n")
            # Limit how many links to read from file and draw
            CIRCOSCONF.write("record_limit  = 10000\n")
            CIRCOSCONF.write("</link>\n")
    CIRCOSCONF.write("</links>\n")
    # Heatmap
    CIRCOSCONF.write("<plots>\n")
    for file in os.listdir(geminiset.pathTMP):
        if "heatmap" in file:
            CIRCOSCONF.write("<plot>\n")
            CIRCOSCONF.write("type    = heatmap\n")
            CIRCOSCONF.write("file    = "+geminiset.pathTMP+"/"+file+"\n")
            CIRCOSCONF.write("color   = "+strLstGradient+"\n")
            CIRCOSCONF.write("r0      = 0.99r\n")
            CIRCOSCONF.write("r1      = 0.92r\n")
            CIRCOSCONF.write("min     = 1\n")
            CIRCOSCONF.write("max     = "+str(len(dicoGBK))+"\n")
            CIRCOSCONF.write("</plot>\n")
    CIRCOSCONF.write("</plots>\n")

    CIRCOSCONF.write("<<include /etc/circos/colors_fonts_patterns.conf>>\n")
    CIRCOSCONF.write("<<include /etc/circos/housekeeping.conf>>\n")
    CIRCOSCONF.close()

    # ***** MAKE PLOT *****#
    print("circos -conf "+geminiset.pathTMP+"/circos.conf")
    os.system(dicoGeminiPath['TOOLS']['circos']+" -conf "+geminiset.pathTMP+"/circos.conf")
